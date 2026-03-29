from __future__ import annotations

import csv
import logging
from typing import Callable, Dict, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from PySide6.QtCore import Qt
from PySide6.QtWidgets import (
    QComboBox,
    QFileDialog,
    QHBoxLayout,
    QLabel,
    QLineEdit,
    QMessageBox,
    QPushButton,
    QSizePolicy,
    QSplitter,
    QTableView,
    QVBoxLayout,
    QWidget,
)

from application.result_display_formatter import format_step_result_row
from ui.results_table_model import ResultsTableModel
from ui.step_log_model import StepLogModel
from ui.workers.results_task_worker import ResultsTaskWorker


log = logging.getLogger(__name__)


class ResultsTab(QWidget):
    def __init__(
        self,
        service,
        run_repo,
        get_project_id: Callable[[], str | None],
        get_base_preset_id: Callable[[], str | None],
        reload_presets_callback: Callable[[str, str | None], None],
        parent=None,
    ):
        super().__init__(parent)
        self.svc = service
        self.run_repo = run_repo
        self.get_project_id = get_project_id
        self.get_base_preset_id = get_base_preset_id
        self.reload_presets_callback = reload_presets_callback
        self._last_results_rows: List[Dict] = []
        self._task_worker: ResultsTaskWorker | None = None
        self._task_generation = 0
        self._busy_action = ""
        self._build_ui()

    def _build_ui(self) -> None:
        layout = QVBoxLayout(self)

        row1 = QHBoxLayout()
        self.run_combo = QComboBox()
        self.btn_refresh_runs = QPushButton("Refresh Runs")
        self.btn_load_results = QPushButton("Load Results")
        self.btn_export_results_csv = QPushButton("Export CSV")
        self.btn_export_results_excel = QPushButton("Export Excel")

        row1.addWidget(QLabel("Run:"))
        row1.addWidget(self.run_combo, 2)
        row1.addWidget(self.btn_refresh_runs)
        row1.addWidget(self.btn_load_results)
        row1.addWidget(self.btn_export_results_csv)
        row1.addWidget(self.btn_export_results_excel)
        layout.addLayout(row1)

        row2 = QHBoxLayout()
        self.result_filter_status = QComboBox()
        self.result_filter_status.addItems(["ALL", "FAIL", "PASS", "SKIP", "ERROR"])

        self.btn_fail_only = QPushButton("FAIL")
        self.btn_error_only = QPushButton("ERROR")
        self.btn_show_all_results = QPushButton("ALL")

        self.result_filter_test_type = QComboBox()
        self.result_filter_test_type.addItem("ALL")
        self.result_filter_band = QComboBox()
        self.result_filter_band.addItem("ALL")
        self.result_filter_standard = QComboBox()
        self.result_filter_standard.addItem("ALL")
        self.result_filter_bw = QComboBox()
        self.result_filter_bw.addItem("ALL")
        self.result_filter_channel = QComboBox()
        self.result_filter_channel.addItem("ALL")

        row2.addWidget(QLabel("Status:"))
        row2.addWidget(self.result_filter_status)
        row2.addWidget(self.btn_show_all_results)
        row2.addWidget(self.btn_fail_only)
        row2.addWidget(self.btn_error_only)

        row2.addSpacing(8)
        row2.addWidget(QLabel("Test:"))
        row2.addWidget(self.result_filter_test_type)

        row2.addSpacing(8)
        row2.addWidget(QLabel("Band:"))
        row2.addWidget(self.result_filter_band)

        row2.addSpacing(8)
        row2.addWidget(QLabel("Std:"))
        row2.addWidget(self.result_filter_standard)

        row2.addSpacing(8)
        row2.addWidget(QLabel("BW:"))
        row2.addWidget(self.result_filter_bw)

        row2.addSpacing(8)
        row2.addWidget(QLabel("CH:"))
        row2.addWidget(self.result_filter_channel)

        layout.addLayout(row2)

        row3 = QHBoxLayout()
        self.result_search = QLineEdit()
        self.result_search.setPlaceholderText("Search test/band/std/ch/bw/reason/key...")
        self.btn_clear_result_filter = QPushButton("Clear Filter")
        self.btn_rerun_from_selection = QPushButton("Re-run from Selection")
        self.lbl_result_summary = QLabel("PASS 0 | FAIL 0 | SKIP 0 | ERROR 0")

        row3.addWidget(QLabel("Search:"))
        row3.addWidget(self.result_search, 2)
        row3.addWidget(self.btn_clear_result_filter)
        row3.addWidget(self.btn_rerun_from_selection)
        row3.addSpacing(12)
        row3.addWidget(self.lbl_result_summary)
        row3.addStretch(1)
        layout.addLayout(row3)

        splitter = QSplitter(Qt.Vertical)

        self.results_table = QTableView()
        self.results_model = ResultsTableModel()
        self.results_table.setModel(self.results_model)
        self.results_table.setSelectionBehavior(QTableView.SelectRows)
        self.results_table.setSelectionMode(QTableView.ExtendedSelection)
        self.results_table.horizontalHeader().setStretchLastSection(True)
        self.results_table.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.results_table.setSortingEnabled(True)
        self.results_table.setAlternatingRowColors(True)
        self.results_table.setStyleSheet("QTableView { alternate-background-color: #141b24; gridline-color: #334155; }")

        self.steps_table = QTableView()
        self.steps_model = StepLogModel()
        self.steps_table.setModel(self.steps_model)
        self.steps_table.setSelectionBehavior(QTableView.SelectRows)
        self.steps_table.horizontalHeader().setStretchLastSection(True)
        self.steps_table.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.steps_table.setAlternatingRowColors(True)
        self.steps_table.setStyleSheet("QTableView { alternate-background-color: #141b24; gridline-color: #334155; }")

        splitter.addWidget(self.results_table)
        splitter.addWidget(self.steps_table)
        splitter.setSizes([700, 300])
        splitter.setStretchFactor(0, 3)
        splitter.setStretchFactor(1, 1)
        layout.addWidget(splitter, 1)

        self.btn_refresh_runs.clicked.connect(self.refresh_runs)
        self.btn_load_results.clicked.connect(self.load_results)
        self.btn_rerun_from_selection.clicked.connect(self.rerun_from_selection)
        self.btn_clear_result_filter.clicked.connect(self.clear_result_filters)
        self.btn_show_all_results.clicked.connect(self.results_show_all)
        self.btn_fail_only.clicked.connect(self.results_fail_only)
        self.btn_error_only.clicked.connect(self.results_error_only)
        self.btn_export_results_csv.clicked.connect(self.export_results_csv)
        self.btn_export_results_excel.clicked.connect(self.export_results_excel)

        self.result_filter_status.currentIndexChanged.connect(self.load_results)
        self.result_filter_test_type.currentIndexChanged.connect(self.load_results)
        self.result_filter_band.currentIndexChanged.connect(self.load_results)
        self.result_filter_standard.currentIndexChanged.connect(self.load_results)
        self.result_filter_bw.currentIndexChanged.connect(self.load_results)
        self.result_filter_channel.currentIndexChanged.connect(self.load_results)
        self.result_search.returnPressed.connect(self.load_results)
        self.results_table.selectionModel().selectionChanged.connect(self.on_result_selection_changed)

    def reset_view(self) -> None:
        self._task_generation += 1
        self._task_worker = None
        self._last_results_rows = []
        self.run_combo.blockSignals(True)
        self.run_combo.clear()
        self.run_combo.blockSignals(False)
        self.results_model.set_rows([])
        self.steps_model.set_rows([])
        self.lbl_result_summary.setText("PASS 0 | FAIL 0 | SKIP 0 | ERROR 0")
        for combo in (
            self.result_filter_status,
            self.result_filter_test_type,
            self.result_filter_band,
            self.result_filter_standard,
            self.result_filter_bw,
            self.result_filter_channel,
        ):
            combo.blockSignals(True)
            if combo.count() > 0:
                combo.setCurrentIndex(0)
            combo.blockSignals(False)
        self.result_search.clear()
        self._update_result_quick_buttons_style()
        self._set_busy(False)

    def _set_busy(self, busy: bool, action: str = "") -> None:
        self._busy_action = action if busy else ""
        self.btn_refresh_runs.setEnabled(not busy)
        self.btn_load_results.setEnabled(not busy)
        self.btn_export_results_csv.setEnabled(not busy)
        self.btn_export_results_excel.setEnabled(not busy)
        self.btn_clear_result_filter.setEnabled(not busy)
        self.btn_rerun_from_selection.setEnabled(not busy)

    def _finish_worker(self, worker: ResultsTaskWorker) -> None:
        if self._task_worker is worker:
            self._task_worker = None
        worker.deleteLater()
        self._set_busy(False)

    def _start_task(self, *, action: str, task, on_success, error_title: str) -> None:
        if self._task_worker is not None and self._task_worker.isRunning():
            log.info("results task skipped | busy_action=%s next_action=%s", self._busy_action, action)
            return

        self._task_generation += 1
        generation = self._task_generation
        worker = ResultsTaskWorker(task)
        self._task_worker = worker
        self._set_busy(True, action=action)

        def _handle_success(payload) -> None:
            self._finish_worker(worker)
            if generation != self._task_generation:
                return
            on_success(payload)

        def _handle_failure(error_text: str) -> None:
            self._finish_worker(worker)
            if generation != self._task_generation:
                return
            log.error("results task failed | action=%s\n%s", action, error_text)
            QMessageBox.critical(self, error_title, error_text)

        worker.succeeded.connect(_handle_success)
        worker.failed.connect(_handle_failure)
        worker.start()

    def _fill_run_combo(self, combo: QComboBox, runs: List[Dict]) -> None:
        current = combo.currentData()
        combo.blockSignals(True)
        combo.clear()
        for r in runs:
            run_id = r.get("run_id", "")
            started_at = r.get("started_at", "")
            status = r.get("status", "")
            preset_name = r.get("preset_name", "") or r.get("preset_id", "")
            equipment_profile = r.get("equipment_profile_name") or ""
            analyzer_name = r.get("analyzer_device_name") or ""
            run_metadata = r.get("run_metadata") or {}
            switch_path = run_metadata.get("switch_path") or ""
            power_control = run_metadata.get("power_control") or {}
            motion_control = run_metadata.get("motion_control") or {}
            extra = ""
            if equipment_profile:
                extra += f" | EQ:{equipment_profile}"
            if analyzer_name:
                extra += f" | AN:{analyzer_name}"
            if switch_path:
                extra += f" | PATH:{switch_path}"
            if power_control.get("enabled"):
                psu_text = "PSU"
                if power_control.get("output_on"):
                    psu_text += ":ON"
                extra += f" | {psu_text}"
            if motion_control.get("enabled"):
                motion_text = "MOTION"
                angle = motion_control.get("turntable_angle_deg")
                height = motion_control.get("mast_height_cm")
                if angle not in (None, "", 0, 0.0):
                    motion_text += f":AZ{angle}"
                if height not in (None, "", 0, 0.0):
                    motion_text += f"/H{height}"
                extra += f" | {motion_text}"
            combo.addItem(f"{started_at} | {status} | {preset_name}{extra}", run_id)

        if current:
            idx = combo.findData(current)
            if idx >= 0:
                combo.setCurrentIndex(idx)
        combo.blockSignals(False)

    def _get_selected_result_filters(self) -> Dict:
        return {
            "status": self.result_filter_status.currentText(),
            "test_type": self.result_filter_test_type.currentText(),
            "band": self.result_filter_band.currentText(),
            "standard": self.result_filter_standard.currentText(),
            "bw_mhz": self.result_filter_bw.currentText(),
            "channel": self.result_filter_channel.currentText(),
            "search": self.result_search.text().strip().lower(),
        }

    def _apply_result_filters(self, rows: List[Dict], filters: Dict) -> List[Dict]:
        filtered = []
        for r in rows:
            if filters["test_type"] != "ALL" and str(r.get("test_type", "")) != filters["test_type"]:
                continue
            if filters["band"] != "ALL" and str(r.get("band", "")) != filters["band"]:
                continue
            if filters["standard"] != "ALL" and str(r.get("standard", "")) != filters["standard"]:
                continue
            if filters["bw_mhz"] != "ALL" and str(r.get("bw_mhz", "")) != filters["bw_mhz"]:
                continue
            if filters["channel"] != "ALL" and str(r.get("channel", "")) != filters["channel"]:
                continue

            search_text = filters["search"]
            if search_text:
                hay = " ".join([
                    str(r.get("test_type", "")),
                    str(r.get("band", "")),
                    str(r.get("standard", "")),
                    str(r.get("group", "")),
                    str(r.get("channel", "")),
                    str(r.get("bw_mhz", "")),
                    str(r.get("reason", "")),
                    str(r.get("test_key", "")),
                ]).lower()
                if search_text not in hay:
                    continue
            filtered.append(r)
        return filtered

    def _refresh_result_filter_options(self, rows: List[Dict]) -> None:
        current_test = self.result_filter_test_type.currentText()
        current_band = self.result_filter_band.currentText()
        current_standard = self.result_filter_standard.currentText()
        current_bw = self.result_filter_bw.currentText()
        current_channel = self.result_filter_channel.currentText()

        test_types = sorted({str(r.get("test_type", "")).strip() for r in rows if r.get("test_type")})
        bands = sorted({str(r.get("band", "")).strip() for r in rows if r.get("band")})
        standards = sorted({str(r.get("standard", "")).strip() for r in rows if r.get("standard")})
        bw_values = sorted({int(r.get("bw_mhz")) for r in rows if r.get("bw_mhz") is not None and str(r.get("bw_mhz")).strip() != ""})
        channel_values = sorted({int(r.get("channel")) for r in rows if r.get("channel") is not None and str(r.get("channel")).strip() != ""})

        combos = [
            self.result_filter_test_type,
            self.result_filter_band,
            self.result_filter_standard,
            self.result_filter_bw,
            self.result_filter_channel,
        ]
        for cb in combos:
            cb.blockSignals(True)

        self.result_filter_test_type.clear()
        self.result_filter_test_type.addItem("ALL")
        for v in test_types:
            self.result_filter_test_type.addItem(v)

        self.result_filter_band.clear()
        self.result_filter_band.addItem("ALL")
        for v in bands:
            self.result_filter_band.addItem(v)

        self.result_filter_standard.clear()
        self.result_filter_standard.addItem("ALL")
        for v in standards:
            self.result_filter_standard.addItem(v)

        self.result_filter_bw.clear()
        self.result_filter_bw.addItem("ALL")
        for v in bw_values:
            self.result_filter_bw.addItem(str(v))

        self.result_filter_channel.clear()
        self.result_filter_channel.addItem("ALL")
        for v in channel_values:
            self.result_filter_channel.addItem(str(v))

        idx = self.result_filter_test_type.findText(current_test)
        self.result_filter_test_type.setCurrentIndex(idx if idx >= 0 else 0)
        idx = self.result_filter_band.findText(current_band)
        self.result_filter_band.setCurrentIndex(idx if idx >= 0 else 0)
        idx = self.result_filter_standard.findText(current_standard)
        self.result_filter_standard.setCurrentIndex(idx if idx >= 0 else 0)
        idx = self.result_filter_bw.findText(current_bw)
        self.result_filter_bw.setCurrentIndex(idx if idx >= 0 else 0)
        idx = self.result_filter_channel.findText(current_channel)
        self.result_filter_channel.setCurrentIndex(idx if idx >= 0 else 0)

        for cb in combos:
            cb.blockSignals(False)

    def _update_results_summary(self, rows: List[Dict]) -> None:
        p = sum(1 for r in rows if (r.get("status") or "").upper() == "PASS")
        f = sum(1 for r in rows if (r.get("status") or "").upper() == "FAIL")
        s = sum(1 for r in rows if (r.get("status") or "").upper() == "SKIP")
        e = sum(1 for r in rows if (r.get("status") or "").upper() == "ERROR")
        self.lbl_result_summary.setText(f"PASS {p} | FAIL {f} | SKIP {s} | ERROR {e}")

    def _update_result_quick_buttons_style(self) -> None:
        base_style = ""
        active_all = "background:#dbeafe; color:#111827; font-weight:600;"
        active_fail = "background:#fee2e2; color:#7f1d1d; font-weight:700;"
        active_error = "background:#ffedd5; color:#7c2d12; font-weight:700;"
        self.btn_show_all_results.setStyleSheet(base_style)
        self.btn_fail_only.setStyleSheet(base_style)
        self.btn_error_only.setStyleSheet(base_style)

        status = self.result_filter_status.currentText()
        if status == "ALL":
            self.btn_show_all_results.setStyleSheet(active_all)
        elif status == "FAIL":
            self.btn_fail_only.setStyleSheet(active_fail)
        elif status == "ERROR":
            self.btn_error_only.setStyleSheet(active_error)

    def refresh_runs(self):
        project_id = self.get_project_id()
        if not project_id:
            QMessageBox.warning(self, "No project", "Select a project.")
            return
        try:
            runs = self.svc.list_runs_for_results(project_id, limit=100)
        except Exception as e:
            log.exception("load runs failed")
            QMessageBox.critical(self, "Load Runs Failed", str(e))
            return
        self._fill_run_combo(self.run_combo, runs)

    def load_results(self):
        project_id = self.get_project_id()
        run_id = self.run_combo.currentData()
        if not project_id or not run_id:
            self.results_model.set_rows([])
            self.steps_model.set_rows([])
            self._update_results_summary([])
            return

        status_filter = self.result_filter_status.currentText()
        active_filters = self._get_selected_result_filters()

        def _task():
            rows = self.svc.get_results_page(
                project_id=project_id,
                run_id=run_id,
                status_filter=status_filter,
                offset=0,
                limit=500,
            )
            return {
                "project_id": project_id,
                "run_id": run_id,
                "rows": rows,
                "filters": active_filters,
            }

        def _apply(payload: Dict) -> None:
            if payload.get("project_id") != self.get_project_id():
                return
            if payload.get("run_id") != self.run_combo.currentData():
                return
            rows = list(payload.get("rows") or [])
            self._last_results_rows = rows
            self._refresh_result_filter_options(rows)
            filtered = self._apply_result_filters(rows, payload.get("filters") or self._get_selected_result_filters())
            self.results_model.set_rows(filtered)
            self.results_table.resizeColumnsToContents()
            self._update_results_summary(filtered)
            self._update_result_quick_buttons_style()

        self._start_task(
            action="load_results",
            task=_task,
            on_success=_apply,
            error_title="Load Results Failed",
        )

    def clear_result_filters(self):
        self.result_filter_status.setCurrentText("ALL")
        self.result_filter_test_type.setCurrentText("ALL")
        self.result_filter_band.setCurrentText("ALL")
        self.result_filter_standard.setCurrentText("ALL")
        self.result_filter_bw.setCurrentText("ALL")
        self.result_filter_channel.setCurrentText("ALL")
        self.result_search.clear()
        self._update_result_quick_buttons_style()
        if self.run_combo.currentData():
            self.load_results()

    def on_result_selection_changed(self, selected, deselected):
        project_id = self.get_project_id()
        if not project_id:
            return
        sel = self.results_table.selectionModel().selectedRows()
        if not sel:
            self.steps_model.set_rows([])
            return

        row = self.results_model.get_row(sel[0].row())
        result_id = row.get("result_id")
        if not result_id:
            self.steps_model.set_rows([])
            return

        try:
            steps = self.run_repo.list_step_results(project_id=project_id, result_id=result_id)
        except Exception:
            log.exception("load step results failed | result_id=%s", result_id)
            self.steps_model.set_rows([])
            return
        self.steps_model.set_rows([format_step_result_row(s) for s in steps])
        self.steps_table.resizeColumnsToContents()

    def rerun_from_selection(self):
        project_id = self.get_project_id()
        if not project_id:
            QMessageBox.information(self, "No project", "Select a project first.")
            return

        base_preset_id = self.get_base_preset_id()
        if not base_preset_id:
            QMessageBox.information(self, "No base preset", "Select a preset (base) first.")
            return

        sel = self.results_table.selectionModel().selectedRows()
        if not sel:
            QMessageBox.information(self, "No selection", "Select result rows first.")
            return

        selected_rows = []
        for idx in sel:
            r = self.results_model.get_row(idx.row())
            if r:
                selected_rows.append(r)

        try:
            new_preset_id = self.svc.create_rerun_preset_from_selected_results(
                project_id=project_id,
                base_preset_id=base_preset_id,
                selected_rows=selected_rows,
            )
            QMessageBox.information(self, "Re-run preset created", f"New preset created.\nPreset ID: {new_preset_id}")
            self.reload_presets_callback(project_id, new_preset_id)
        except Exception as e:
            log.exception("rerun from selection failed")
            QMessageBox.warning(self, "Re-run failed", str(e))

    def _fetch_results_for_export(self, *, project_id: str, run_id: str, status_filter: str, limit: int = 20000) -> List[Dict]:
        return self.svc.get_results_page(
            project_id=project_id,
            run_id=run_id,
            status_filter=status_filter,
            offset=0,
            limit=limit,
        )

    def _write_results_csv(self, path: str, rows: List[Dict]) -> None:
        cols = [
            ("status", "Status"),
            ("test_type", "Test"),
            ("band", "Band"),
            ("standard", "Standard"),
            ("group", "Group"),
            ("channel", "CH"),
            ("bw_mhz", "BW(MHz)"),
            ("margin_db", "Margin(dB)"),
            ("measured_value", "Measured"),
            ("limit_value", "Limit"),
            ("reason", "Reason"),
            ("test_key", "Key"),
            ("result_id", "Result ID"),
        ]
        with open(path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerow([header for _, header in cols])
            for row in rows:
                writer.writerow([row.get(key, "") for key, _ in cols])

    def _write_results_excel(self, path: str, rows: List[Dict]) -> None:
        cols = [
            ("status", "Status"),
            ("test_type", "Test"),
            ("band", "Band"),
            ("standard", "Standard"),
            ("group", "Group"),
            ("channel", "CH"),
            ("bw_mhz", "BW(MHz)"),
            ("margin_db", "Margin(dB)"),
            ("measured_value", "Measured"),
            ("limit_value", "Limit"),
            ("reason", "Reason"),
            ("test_key", "Key"),
            ("result_id", "Result ID"),
        ]
        wb = Workbook()
        ws = wb.active
        ws.title = "Results"
        header_font = Font(bold=True)
        for c, (_, header) in enumerate(cols, start=1):
            cell = ws.cell(row=1, column=c, value=header)
            cell.font = header_font
            cell.alignment = Alignment(vertical="center")

        for r_i, row in enumerate(rows, start=2):
            for c_i, (key, _) in enumerate(cols, start=1):
                ws.cell(row=r_i, column=c_i, value=row.get(key, ""))

        for c_i in range(1, len(cols) + 1):
            ws.column_dimensions[ws.cell(row=1, column=c_i).column_letter].width = 16
        ws.column_dimensions["K"].width = 40
        wb.save(path)

    def export_results_csv(self):
        project_id = self.get_project_id()
        run_id = self.run_combo.currentData()
        status_filter = self.result_filter_status.currentText()
        if not project_id or not run_id:
            QMessageBox.information(self, "Export CSV", "Select a run first.")
            return

        path, _ = QFileDialog.getSaveFileName(self, "Export Results (CSV)", "results.csv", "CSV (*.csv)")
        if not path:
            return

        def _task():
            rows = self._fetch_results_for_export(
                project_id=project_id,
                run_id=run_id,
                status_filter=status_filter,
                limit=20000,
            )
            if not rows:
                return {"path": path, "rows": []}
            self._write_results_csv(path, rows)
            return {"path": path, "rows": rows}

        def _done(payload: Dict) -> None:
            rows = list(payload.get("rows") or [])
            if not rows:
                QMessageBox.information(self, "Export CSV", "No rows to export.")
                return
            QMessageBox.information(self, "Export CSV", f"Saved:\n{payload.get('path', path)}")

        self._start_task(
            action="export_csv",
            task=_task,
            on_success=_done,
            error_title="Export CSV Failed",
        )

    def export_results_excel(self):
        project_id = self.get_project_id()
        run_id = self.run_combo.currentData()
        status_filter = self.result_filter_status.currentText()
        if not project_id or not run_id:
            QMessageBox.information(self, "Export Excel", "Select a run first.")
            return

        path, _ = QFileDialog.getSaveFileName(self, "Export Results (Excel)", "results.xlsx", "Excel (*.xlsx)")
        if not path:
            return

        def _task():
            rows = self._fetch_results_for_export(
                project_id=project_id,
                run_id=run_id,
                status_filter=status_filter,
                limit=20000,
            )
            if not rows:
                return {"path": path, "rows": []}
            self._write_results_excel(path, rows)
            return {"path": path, "rows": rows}

        def _done(payload: Dict) -> None:
            rows = list(payload.get("rows") or [])
            if not rows:
                QMessageBox.information(self, "Export Excel", "No rows to export.")
                return
            QMessageBox.information(self, "Export Excel", f"Saved:\n{payload.get('path', path)}")

        self._start_task(
            action="export_excel",
            task=_task,
            on_success=_done,
            error_title="Export Excel Failed",
        )

    def results_show_all(self):
        self.result_filter_status.setCurrentText("ALL")
        self._update_result_quick_buttons_style()
        self.load_results()

    def results_fail_only(self):
        self.result_filter_status.setCurrentText("FAIL")
        self._update_result_quick_buttons_style()
        self.load_results()

    def results_error_only(self):
        self.result_filter_status.setCurrentText("ERROR")
        self._update_result_quick_buttons_style()
        self.load_results()
