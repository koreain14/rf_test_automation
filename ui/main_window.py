from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Optional

from PySide6.QtCore import Qt, QThread, Signal
from PySide6.QtGui import QAction, QStandardItem, QStandardItemModel
from PySide6.QtWidgets import (
    QComboBox, QHBoxLayout, QLabel, QMainWindow, QMessageBox, QPushButton,
    QSplitter, QToolBar, QTreeView, QTableView, QVBoxLayout, QWidget, QTabWidget, QListWidget, QLineEdit, QSizePolicy, QProgressBar
)

from application.plan_service import PlanService
from domain.models import OverrideRule, Preset, Recipe, RuleSet
from ui.table_model import CaseTableModel
from ui.results_table_model import ResultsTableModel
from ui.execution_order_dialog import ExecutionOrderDialog
from ui.step_log_model import StepLogModel
import uuid

import json
from datetime import datetime
from PySide6.QtWidgets import QFileDialog

import csv
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment


@dataclass
class PlanContext:
    project_id: str
    preset_id: str
    ruleset: RuleSet
    preset: Preset
    recipe: Recipe
    overrides: List[OverrideRule]


class MainWindow(QMainWindow):
    PAGE_SIZE = 200

    def __init__(self, plan_service: PlanService, run_repo, run_service):
        super().__init__()
        
        self._result_rows = []
        self.svc = plan_service
        self._plans = {}
        
        self.run_repo = run_repo
        self.run_service = run_service

        self.setWindowTitle("RF Test Platform (Prototype)")
        self.resize(1200, 800)

        self.project_id: Optional[str] = None
        self.preset_id: Optional[str] = None

        self._plans: Dict[str, PlanContext] = {}  # plan_node_id -> PlanContext
        self._current_plan_node_id: Optional[str] = None
        self._current_filter: Optional[Dict[str, Any]] = None
        self._current_offset: int = 0

        self._build_ui()
        self._load_initial_data()
        
       
        self._worker = None
        self._last_run_id = None
        self._last_results_rows: list[dict] = [] # Check
        
        self._scenario_worker = None
        self._scenario_total_cases = 0
        self._scenario_processed_cases = 0
        self._scenario_run_summaries = []
        
        self._run_total_cases = 0
        self._run_processed_cases = 0
        
        self._run_pass_count = 0
        self._run_fail_count = 0
        self._run_skip_count = 0
        self._run_error_count = 0
        self._running_preset_name = ""
        

    def _build_ui(self):
        toolbar = QToolBar("Main")
        tabs = QTabWidget()
        self.btn_order = QPushButton("Execution Order")

        self.addToolBar(toolbar)

        self.project_combo = QComboBox()
        self.preset_combo = QComboBox()
        
        # 실행 버튼
        self.btn_start = QPushButton("Start Run")
        self.btn_run_scenario = QPushButton("Run Scenario")
        self.btn_stop = QPushButton("Stop")
        self.btn_rerun = QPushButton("Create Re-run (FAIL)")
        

        self.lbl_status = QLabel("Idle")
        
        self.progress_run = QProgressBar()
        self.progress_run.setMinimum(0)
        self.progress_run.setMaximum(100)
        self.progress_run.setValue(0)
        self.progress_run.setTextVisible(True)
        self.progress_run.setFormat("Idle")
        self.progress_run.setFixedWidth(220)

        toolbar.addSeparator()
        toolbar.addWidget(self.btn_start)
        toolbar.addWidget(self.btn_run_scenario)
        toolbar.addWidget(self.btn_stop)
        toolbar.addWidget(self.btn_rerun)
        toolbar.addWidget(self.lbl_status)
        toolbar.addWidget(self.progress_run)

        toolbar.addSeparator()
        toolbar.addWidget(self.btn_order)
        self.btn_order.clicked.connect(self.on_edit_execution_order)


        self.btn_start.clicked.connect(self.on_start_run)
        self.btn_run_scenario.clicked.connect(self.on_start_scenario_run)
        self.btn_stop.clicked.connect(self.on_stop_run)
        self.btn_rerun.clicked.connect(self.on_create_rerun)

        toolbar.addWidget(QLabel(" Project: "))
        toolbar.addWidget(self.project_combo)
        toolbar.addSeparator()
        toolbar.addWidget(QLabel(" Preset: "))
        toolbar.addWidget(self.preset_combo)

        self.btn_add_plan = QPushButton("Add Plan")
        self.btn_remove_plan = QPushButton("Remove Plan")
        self.btn_reload = QPushButton("Reload Plan")
        self.btn_more = QPushButton("Load More")
        self.btn_skip = QPushButton("Skip Selected")

        toolbar.addSeparator()
        toolbar.addWidget(self.btn_add_plan)
        toolbar.addWidget(self.btn_remove_plan)
        toolbar.addWidget(self.btn_more)
        toolbar.addWidget(self.btn_skip)
        toolbar.addWidget(self.btn_reload)

        self.btn_add_plan.clicked.connect(self.on_add_plan)
        self.btn_remove_plan.clicked.connect(self.on_remove_plan_from_scenario)
        self.btn_more.clicked.connect(self.on_load_more)
        self.btn_skip.clicked.connect(self.on_skip_selected)
        self.btn_reload.clicked.connect(self.on_reload_plan)

        self.project_combo.currentIndexChanged.connect(self.on_project_changed)
        self.preset_combo.currentIndexChanged.connect(self.on_preset_changed)

        plan_splitter = QSplitter(Qt.Horizontal)

        # Tree
        self.tree = QTreeView()
        self.tree_model = QStandardItemModel()
        self.tree_model.setHorizontalHeaderLabels(["Scenario Tree"])
        self.tree.setModel(self.tree_model)
        self.tree.clicked.connect(self.on_tree_clicked)

        # Table
        self.table = QTableView()
        self.case_model = CaseTableModel()
        self.table.setModel(self.case_model)
        self.table.setSelectionBehavior(QTableView.SelectRows)
        self.table.setSelectionMode(QTableView.ExtendedSelection)
        self.table.horizontalHeader().setStretchLastSection(True)

        plan_splitter.addWidget(self.tree)
        plan_splitter.addWidget(self.table)
        # plan_splitter.setSizes([350, 850])
        
        self.tree.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.table.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        
        # (2) ✅ 리사이즈 비율(트리:테이블 = 1:3 추천)
        plan_splitter.setStretchFactor(0, 1)
        plan_splitter.setStretchFactor(1, 3)

        # (3) ✅ Plan 탭 컨테이너 + 레이아웃
        plan_page = QWidget()
        plan_layout = QVBoxLayout(plan_page)
        plan_layout.setContentsMargins(6, 6, 6, 6)
        plan_layout.setSpacing(6)

        # (4) ✅ Scenario 버튼 row (toolbar 대신 Plan 탭 상단에)
        scenario_bar = QHBoxLayout()
        scenario_bar.setSpacing(6)
        
        self.btn_save_scenario = QPushButton("Save Scenario")
        self.btn_load_scenario = QPushButton("Load Scenario")
        self.btn_clear_scenario = QPushButton("Clear Scenario")

        self.btn_save_scenario.clicked.connect(self.on_save_scenario)
        self.btn_load_scenario.clicked.connect(self.on_load_scenario)
        self.btn_clear_scenario.clicked.connect(self.on_clear_scenario)

        scenario_bar.addWidget(self.btn_save_scenario)
        scenario_bar.addWidget(self.btn_load_scenario)
        scenario_bar.addWidget(self.btn_clear_scenario)
        scenario_bar.addStretch(1)  # ✅ 왼쪽 버튼 고정 + 오른쪽 여백 유연하게

        # (5) ✅ 레이아웃에 추가: 상단 row + splitter(남는 공간 전부 사용)
        plan_layout.addLayout(scenario_bar)
        plan_layout.addWidget(plan_splitter, 1)

        # (6) ✅ 탭에 추가 (기존 tabs.addTab(plan_splitter, "Plan") 대신)
        tabs.addTab(plan_page, "Plan")
            
       
        # --- Results tab (새로 만들 위젯) ---
        self.results_widget = self._build_results_tab()
        tabs.addTab(self.results_widget, "Results")
        
        

        self.setCentralWidget(tabs)
        
    def _build_results_tab(self) -> QWidget:
        w = QWidget()
        layout = QVBoxLayout(w)

        # -----------------------------
        # Row 1: Run controls
        # -----------------------------
        row1 = QHBoxLayout()
        self.run_combo = QComboBox()
        self.btn_refresh_runs = QPushButton("Refresh Runs")
        self.btn_load_results = QPushButton("Load Results")

        row1.addWidget(QLabel("Run:"))
        row1.addWidget(self.run_combo, 2)
        row1.addWidget(self.btn_refresh_runs)
        row1.addWidget(self.btn_load_results)

        layout.addLayout(row1)

        # -----------------------------
        # Row 2: Status quick buttons + filters
        # -----------------------------
        row2 = QHBoxLayout()

        self.result_filter_status = QComboBox()
        self.result_filter_status.addItems(["ALL", "FAIL", "PASS", "SKIP", "ERROR"])

        self.btn_fail_only = QPushButton("FAIL")
        self.btn_error_only = QPushButton("ERROR")
        self.btn_show_all_results = QPushButton("ALL")

        self.result_filter_test_type = QComboBox()
        self.result_filter_test_type.addItem("ALL")

        self.result_filter_band = QComboBox()
        self.result_filter_band.addItem("ALL")

        self.result_filter_standard = QComboBox()
        self.result_filter_standard.addItem("ALL")

        self.result_filter_bw = QComboBox()
        self.result_filter_bw.addItem("ALL")

        self.result_filter_channel = QComboBox()
        self.result_filter_channel.addItem("ALL")

        row2.addWidget(QLabel("Status:"))
        row2.addWidget(self.result_filter_status)
        row2.addWidget(self.btn_show_all_results)
        row2.addWidget(self.btn_fail_only)
        row2.addWidget(self.btn_error_only)

        row2.addSpacing(8)
        row2.addWidget(QLabel("Test:"))
        row2.addWidget(self.result_filter_test_type)

        row2.addSpacing(8)
        row2.addWidget(QLabel("Band:"))
        row2.addWidget(self.result_filter_band)

        row2.addSpacing(8)
        row2.addWidget(QLabel("Std:"))
        row2.addWidget(self.result_filter_standard)

        row2.addSpacing(8)
        row2.addWidget(QLabel("BW:"))
        row2.addWidget(self.result_filter_bw)

        row2.addSpacing(8)
        row2.addWidget(QLabel("CH:"))
        row2.addWidget(self.result_filter_channel)

        layout.addLayout(row2)

        # -----------------------------
        # Row 3: Search / Clear / Re-run / Summary
        # -----------------------------
        row3 = QHBoxLayout()

        self.result_search = QLineEdit()
        self.result_search.setPlaceholderText("Search test/band/std/ch/bw/reason/key...")

        self.btn_clear_result_filter = QPushButton("Clear Filter")
        self.btn_rerun_from_selection = QPushButton("Re-run from Selection")
        self.lbl_result_summary = QLabel("PASS 0 | FAIL 0 | SKIP 0 | ERROR 0")

        row3.addWidget(QLabel("Search:"))
        row3.addWidget(self.result_search, 2)
        row3.addWidget(self.btn_clear_result_filter)
        row3.addWidget(self.btn_rerun_from_selection)
        row3.addSpacing(12)
        row3.addWidget(self.lbl_result_summary)
        row3.addStretch(1)

        layout.addLayout(row3)

        # -----------------------------
        # Tables
        # -----------------------------
        splitter = QSplitter(Qt.Vertical)

        self.results_table = QTableView()
        self.results_model = ResultsTableModel()
        self.results_table.setModel(self.results_model)
        self.results_table.setSelectionBehavior(QTableView.SelectRows)
        self.results_table.setSelectionMode(QTableView.ExtendedSelection)
        self.results_table.horizontalHeader().setStretchLastSection(True)
        self.results_table.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.results_table.setSortingEnabled(True)   # 정렬 가능

        self.steps_table = QTableView()
        self.steps_model = StepLogModel()
        self.steps_table.setModel(self.steps_model)
        self.steps_table.setSelectionBehavior(QTableView.SelectRows)
        self.steps_table.horizontalHeader().setStretchLastSection(True)
        self.steps_table.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

        splitter.addWidget(self.results_table)
        splitter.addWidget(self.steps_table)
        splitter.setSizes([700, 300])
        splitter.setStretchFactor(0, 3)
        splitter.setStretchFactor(1, 1)

        layout.addWidget(splitter, 1)

        # -----------------------------
        # Signals
        # -----------------------------
        self.btn_refresh_runs.clicked.connect(self.on_refresh_runs)
        self.btn_load_results.clicked.connect(self.on_load_results)
        self.btn_rerun_from_selection.clicked.connect(self.on_rerun_from_selection)
        self.btn_clear_result_filter.clicked.connect(self.on_clear_result_filters)

        self.btn_show_all_results.clicked.connect(self.on_results_show_all)
        self.btn_fail_only.clicked.connect(self.on_results_fail_only)
        self.btn_error_only.clicked.connect(self.on_results_error_only)

        self.result_filter_status.currentIndexChanged.connect(self.on_load_results)
        self.result_filter_test_type.currentIndexChanged.connect(self.on_load_results)
        self.result_filter_band.currentIndexChanged.connect(self.on_load_results)
        self.result_filter_standard.currentIndexChanged.connect(self.on_load_results)
        self.result_filter_bw.currentIndexChanged.connect(self.on_load_results)
        self.result_filter_channel.currentIndexChanged.connect(self.on_load_results)

        self.result_search.returnPressed.connect(self.on_load_results)

        self.results_table.selectionModel().selectionChanged.connect(self.on_result_selection_changed)

        return w

    def _load_initial_data(self):
        project_id, preset_id = self.svc.ensure_demo_project_and_preset()
        self._reload_projects(select_project_id=project_id)
        self._reload_presets(project_id, select_preset_id=preset_id)

    # ✅ 여기 추가: 콤보 상태를 강제로 내부 변수에 반영
        self.project_id = self.project_combo.currentData()
        self.preset_id = self.preset_combo.currentData()

    # ✅ 그리고 트리/테이블 리셋 등 기존 로직을 태워야 하면 직접 호출
        if self.project_id:
            self.on_project_changed(self.project_combo.currentIndex())
        if self.preset_id:
            self.on_preset_changed(self.preset_combo.currentIndex())
    
    def _reload_projects(self, select_project_id: Optional[str] = None):
        self.project_combo.blockSignals(True)
        self.project_combo.clear()

        projects = self.svc.list_projects()
        for p in projects:
            self.project_combo.addItem(p["name"], userData=p["project_id"])

        self.project_combo.blockSignals(False)

        if select_project_id:
            idx = self.project_combo.findData(select_project_id)
            if idx >= 0:
                self.project_combo.setCurrentIndex(idx)

    def _reload_presets(self, project_id: str, select_preset_id: Optional[str] = None):
        self.preset_combo.blockSignals(True)
        self.preset_combo.clear()

        presets = self.svc.list_presets(project_id)
        for pr in presets:
            self.preset_combo.addItem(pr["name"], userData=pr["preset_id"])

        self.preset_combo.blockSignals(False)

        if select_preset_id:
            idx = self.preset_combo.findData(select_preset_id)
            if idx >= 0:
                self.preset_combo.setCurrentIndex(idx)

    def on_reload_plan(self):
        if not self._current_plan_node_id:
            return
        ctx = self._plans.get(self._current_plan_node_id)
        if not ctx:
            return

        # DB에서 overrides 다시 로드
        ctx.overrides = self.svc.load_override_objs(ctx.preset_id)

        # 테이블/페이징 리셋 후 다시 로딩
        self._current_offset = 0
        self.case_model.clear()
        self._load_page()
        
    # ---------- events ----------
    def on_project_changed(self, _idx: int):
        pid = self.project_combo.currentData()
        if not pid:
            return
        self.project_id = pid
        self._reload_presets(pid)
        # 트리/테이블 reset
        self.tree_model.removeRows(0, self.tree_model.rowCount())
        self.case_model.clear()

    def on_preset_changed(self, _idx: int):
        self.preset_id = self.preset_combo.currentData()

    def on_add_plan(self):
        import logging
        logger = logging.getLogger(__name__)
        if not self.project_id:
            QMessageBox.warning(self, "No project", "Select a project.")
            return
        if not self.preset_id:
            QMessageBox.warning(self, "No preset", "Select a preset.")
            return

        try:
            ruleset, preset, recipe, overrides = self.svc.build_recipe_from_preset(self.preset_id)
        except Exception as e:
            logger.exception("build_recipe_from_preset failed")
            QMessageBox.critical(self, "Add plan failed", str(e))
            return

    # 이하 동일...

        # Plan node id(트리에서 찾기 위한 임의 키)
        plan_id = f"PLAN::{uuid.uuid4()}"   
        
        ctx = PlanContext(
            project_id=self.project_id,
            preset_id=self.preset_id,
            ruleset=ruleset,
            preset=preset,
            recipe=recipe,
            overrides=overrides
        )
        self._plans[plan_id] = ctx

        # 트리 표시: Plan > test_type 그룹
        root = self.tree_model.invisibleRootItem()
        plan_item = QStandardItem(f"{preset.name}  ({recipe.band}/{recipe.standard}/{recipe.plan_mode})")
        plan_item.setData(plan_id, role=Qt.UserRole)
        root.appendRow(plan_item)

        for t in recipe.test_types:
            child = QStandardItem(f"{t}")
            child.setData(plan_id, role=Qt.UserRole)  # 같은 plan_node_id
            child.setData({"test_type": t}, role=Qt.UserRole + 1)  # filter
            plan_item.appendRow(child)

        self.tree.expand(plan_item.index())
        self.tree.setCurrentIndex(plan_item.index())
        self._select_tree_node(plan_item)

    def _current_plan_id(self) -> str | None:
        idx = self.tree.currentIndex()
        if not idx.isValid():
            return None
        item = self.tree_model.itemFromIndex(idx)
        if item is None:
            return None
        plan_id = item.data(Qt.UserRole)
        return str(plan_id) if plan_id else None
    
    def _remove_plan_item_from_tree(self, plan_id: str) -> bool:
        root = self.tree_model.invisibleRootItem()
        for row in range(root.rowCount()):
            it = root.child(row)
            if it and str(it.data(Qt.UserRole)) == str(plan_id):
                root.removeRow(row)
                return True
        return False
    
    def on_remove_plan_from_scenario(self):
        plan_id = self._current_plan_id()
        if not plan_id or plan_id not in self._plans:
            QMessageBox.information(self, "Remove Plan", "Select a plan (or a test under a plan).")
            return

        ctx = self._plans[plan_id]
        ret = QMessageBox.question(
            self,
            "Remove Plan",
            f"Remove this plan from the scenario?\n\n{ctx.preset.name} ({ctx.recipe.band}/{ctx.recipe.standard}/{ctx.recipe.plan_mode})",
            QMessageBox.Yes | QMessageBox.No
        )
        if ret != QMessageBox.Yes:
            return

        # 1) 메모리에서 제거
        self._plans.pop(plan_id, None)

        # 2) 트리에서 제거
        self._remove_plan_item_from_tree(plan_id)

        # 3) 우측 화면 초기화(있으면)
        if hasattr(self, "_clear_cases_view"):
            self._clear_cases_view()
            
        # remove한 plan이 현재 선택 plan이면 상태 리셋
        if self._current_plan_node_id == plan_id:
           self._current_plan_node_id = None
           self._current_filter = None
           self._current_offset = 0
           self.case_model.clear()
        
    def on_tree_clicked(self, index):
        item = self.tree_model.itemFromIndex(index)
        if item:
            self._select_tree_node(item)

    def _select_tree_node(self, item: QStandardItem):
        plan_node_id = item.data(Qt.UserRole)
        if not plan_node_id or plan_node_id not in self._plans:
            return

        self._current_plan_node_id = plan_node_id

        # filter info (child node only)
        filter_ = item.data(Qt.UserRole + 1)
        self._current_filter = filter_ if isinstance(filter_, dict) else None

        # reset paging
        self._current_offset = 0
        self.case_model.clear()

        self._load_page()

    def _load_page(self):
        if not self._current_plan_node_id:
            return
        ctx = self._plans[self._current_plan_node_id]

        rows = self.svc.get_cases_page(
            ruleset=ctx.ruleset,
            recipe=ctx.recipe,
            overrides=ctx.overrides,
            filter_=self._current_filter,
            offset=self._current_offset,
            limit=self.PAGE_SIZE
        )
        self.case_model.append_rows(rows)
        self._current_offset += len(rows)

        # 컬럼 폭 간단 조정
        self.table.resizeColumnsToContents()

    def on_load_more(self):
        self._load_page()

    def on_skip_selected(self):
        if not self._current_plan_node_id:
            return
        ctx = self._plans[self._current_plan_node_id]

        sel = self.table.selectionModel().selectedRows()
        cases = []
        for idx in sel:
            c = self.case_model.get_case(idx.row())
            if c:
                cases.append(c)

        try:
            self.svc.create_skip_override_for_selection(
                project_id=ctx.project_id,
                preset_id=ctx.preset_id,
                cases=cases,
                priority=100
            )
            QMessageBox.information(self, "Override created", f"Created 1 grouped skip override for {len(cases)} cases.")
        except Exception as e:
            # fallback: 기존 방식(개별 생성)
            created = 0
            for c in cases:
                self.svc.create_skip_override_for_case(ctx.project_id, ctx.preset_id, c, priority=100)
                created += 1
            QMessageBox.warning(self, "Grouped skip failed",
                                f"{e}\n\nFallback: created {created} individual overrides.")
        self.on_reload_plan()
        
    def _on_run_progress(self, count: int, last_status: str):
        self._run_processed_cases = count

        # 마지막 status 기준 누적 카운트 반영
        st = (last_status or "").upper()
        if st == "PASS":
            self._run_pass_count += 1
        elif st == "FAIL":
            self._run_fail_count += 1
        elif st == "SKIP":
            self._run_skip_count += 1
        elif st == "ERROR":
            self._run_error_count += 1

        total = self._run_total_cases if self._run_total_cases > 0 else 0

        # progress bar
        if total > 0:
            self.progress_run.setMaximum(total)
            self.progress_run.setValue(min(count, total))
            self.progress_run.setFormat(f"{count} / {total}")
        else:
            self.progress_run.setMaximum(100)
            self.progress_run.setValue(0)
            self.progress_run.setFormat(f"{count}")

        run_short = self._last_run_id[:8] if self._last_run_id else "--------"

        # 상태 라벨: 진행 수 + 요약 카운트 + 마지막 결과
        if total > 0:
            self.lbl_status.setText(
                f"RUNNING {run_short} | {count}/{total} | "
                f"P:{self._run_pass_count} F:{self._run_fail_count} "
                f"S:{self._run_skip_count} E:{self._run_error_count} | "
                f"last={st}"
            )
        else:
            self.lbl_status.setText(
                f"RUNNING {run_short} | {count} | "
                f"P:{self._run_pass_count} F:{self._run_fail_count} "
                f"S:{self._run_skip_count} E:{self._run_error_count} | "
                f"last={st}"
            )
    
    def _on_scenario_run_progress(self, processed: int, total: int, preset_name: str, last_status: str):
        self._scenario_processed_cases = processed

        if total > 0:
            self.progress_run.setMaximum(total)
            self.progress_run.setValue(min(processed, total))
            self.progress_run.setFormat(f"{processed} / {total}")
        else:
            self.progress_run.setMaximum(100)
            self.progress_run.setValue(0)
            self.progress_run.setFormat(str(processed))

        self.lbl_status.setText(
            f"SCENARIO RUN | {processed}/{total} | {preset_name} | last={last_status}"
        )
            
    def on_start_run(self):
        if not self._current_plan_node_id:
            QMessageBox.information(self, "No plan", "Add a plan and select it in the tree first.")
            return
        if self._worker and self._worker.isRunning():
            QMessageBox.information(self, "Running", "A run is already in progress.")
            return

        ctx = self._plans[self._current_plan_node_id]

        self._running_preset_name = ctx.preset.name
        self._run_pass_count = 0
        self._run_fail_count = 0
        self._run_skip_count = 0
        self._run_error_count = 0

        # 1) 전체 케이스 수 계산
        try:
            total_cases = self.svc.count_cases(
                ruleset=ctx.ruleset,
                recipe=ctx.recipe,
                overrides=ctx.overrides,
                filter_=self._current_filter,
            )
        except Exception:
            total_cases = 0

        self._run_total_cases = total_cases
        self._run_processed_cases = 0

        # 2) progress bar 초기화
        if total_cases > 0:
            self.progress_run.setMaximum(total_cases)
            self.progress_run.setValue(0)
            self.progress_run.setFormat(f"0 / {total_cases}")
        else:
            self.progress_run.setMaximum(100)
            self.progress_run.setValue(0)
            self.progress_run.setFormat("0")

        # run 생성
        run_id = self.run_repo.create_run(ctx.project_id, ctx.preset_id)
        self._last_run_id = run_id
        self.lbl_status.setText(f"RUNNING {run_id[:8]} ...")

        # worker 시작
        self._worker = RunWorker(
            run_service=self.run_service,
            project_id=ctx.project_id,
            preset_id=ctx.preset_id,
            run_id=run_id,
            ruleset=ctx.ruleset,
            recipe=ctx.recipe,
            overrides=ctx.overrides,
        )
        self._worker.progress.connect(self._on_run_progress)
        self._worker.finished.connect(self._on_run_finished)
        self._worker.start()
        
    def on_start_scenario_run(self):
        if self._worker and self._worker.isRunning():
            QMessageBox.information(self, "Running", "A single-plan run is already in progress.")
            return
        if self._scenario_worker and self._scenario_worker.isRunning():
            QMessageBox.information(self, "Running", "A scenario run is already in progress.")
            return

        plan_ids = self._scenario_plan_ids_in_tree_order()
        if not plan_ids:
            QMessageBox.information(self, "No plans", "Add plans to the scenario first.")
            return

        plan_snapshots = []
        total_cases = 0

        for plan_id in plan_ids:
            ctx = self._plans.get(plan_id)
            if not ctx:
                continue

            try:
                cnt = self.svc.count_cases(
                    ruleset=ctx.ruleset,
                    recipe=ctx.recipe,
                    overrides=ctx.overrides,
                    filter_=None,
                )
            except Exception:
                cnt = 0

            total_cases += cnt

            plan_snapshots.append({
                "plan_id": plan_id,
                "project_id": ctx.project_id,
                "preset_id": ctx.preset_id,
                "preset_name": ctx.preset.name,
                "ruleset": ctx.ruleset,
                "recipe": ctx.recipe,
                "overrides": ctx.overrides,
                "case_count": cnt,
            })

        if not plan_snapshots:
            QMessageBox.information(self, "No plans", "No valid plans found in the scenario.")
            return

        self._scenario_total_cases = total_cases
        self._scenario_processed_cases = 0
        self._scenario_run_summaries = []

        self.progress_run.setMaximum(total_cases if total_cases > 0 else 100)
        self.progress_run.setValue(0)
        self.progress_run.setFormat(f"0 / {total_cases}" if total_cases > 0 else "0")

        self.lbl_status.setText(f"SCENARIO RUN | 0/{total_cases}")

        self._scenario_worker = ScenarioRunWorker(
            run_service=self.run_service,
            run_repo=self.run_repo,
            plan_snapshots=plan_snapshots,
            total_cases=total_cases,
        )
        self._scenario_worker.progress.connect(self._on_scenario_run_progress)
        self._scenario_worker.finished.connect(self._on_scenario_run_finished)
        self._scenario_worker.start()
        
    def _on_run_finished(self, final_status: str, run_id: str, error_text: str):
        self.run_repo.finish_run(run_id, final_status)

        if self._run_total_cases > 0:
            self.progress_run.setMaximum(self._run_total_cases)
            self.progress_run.setValue(self._run_total_cases)
            self.progress_run.setFormat(f"{self._run_total_cases} / {self._run_total_cases}")
        else:
            self.progress_run.setMaximum(100)
            self.progress_run.setValue(100)
            self.progress_run.setFormat("Done")

        # 현재 plan에서 preset 이름 가져오기
        preset_name = self._running_preset_name or "Unknown"

        # 결과 상태 개수 집계
        try:
            counts = self.run_repo.get_run_status_counts(
                project_id=self.project_id,
                run_id=run_id,
            )
        except Exception:
            counts = {"PASS": 0, "FAIL": 0, "SKIP": 0, "ERROR": 0}

        pass_cnt = counts.get("PASS", 0)
        fail_cnt = counts.get("FAIL", 0)
        skip_cnt = counts.get("SKIP", 0)
        error_cnt = counts.get("ERROR", 0)

        if final_status == "ERROR":
            self.lbl_status.setText(f"ERROR {run_id[:8]}")
            msg = (
                f"Run finished\n\n"
                f"Preset : {preset_name}\n"
                f"Status : {final_status}\n"
                f"PASS   : {pass_cnt}\n"
                f"FAIL   : {fail_cnt}\n"
                f"SKIP   : {skip_cnt}\n"
                f"ERROR  : {error_cnt}\n\n"
                f"{error_text or 'Unknown error'}"
            )
            QMessageBox.critical(self, "Run ERROR", msg)
            return

        self.lbl_status.setText(
            f"{final_status} {run_id[:8]} | "
            f"P:{pass_cnt} F:{fail_cnt} S:{skip_cnt} E:{error_cnt}"
        )

        msg = (
            f"Run completed\n\n"
            f"Preset : {preset_name}\n"
            f"Status : {final_status}\n"
            f"PASS   : {pass_cnt}\n"
            f"FAIL   : {fail_cnt}\n"
            f"SKIP   : {skip_cnt}\n"
            f"ERROR  : {error_cnt}"
        )

        QMessageBox.information(self, "Run finished", msg)

        # runs 콤보 자동 갱신
        self.on_refresh_runs()
        
    def _on_scenario_run_finished(self, final_status: str, summaries: list, error_text: str):
        self._scenario_run_summaries = summaries or []

        total = self._scenario_total_cases
        if total > 0:
            self.progress_run.setMaximum(total)
            self.progress_run.setValue(total)
            self.progress_run.setFormat(f"{total} / {total}")
        else:
            self.progress_run.setMaximum(100)
            self.progress_run.setValue(100)
            self.progress_run.setFormat("Done")

        lines = ["Scenario completed", ""]

        total_pass = 0
        total_fail = 0
        total_skip = 0
        total_error = 0

        for s in self._scenario_run_summaries:
            preset_name = s.get("preset_name", "Unknown")
            run_status = s.get("final_status", "")
            counts = s.get("counts", {}) or {}

            p = counts.get("PASS", 0)
            f = counts.get("FAIL", 0)
            sk = counts.get("SKIP", 0)
            e = counts.get("ERROR", 0)

            total_pass += p
            total_fail += f
            total_skip += sk
            total_error += e

            lines.append(
                f"{preset_name} | {run_status} | P:{p} F:{f} S:{sk} E:{e}"
            )

        lines.append("")
        lines.append(f"TOTAL | P:{total_pass} F:{total_fail} S:{total_skip} E:{total_error}")

        if final_status == "ERROR":
            self.lbl_status.setText(
                f"SCENARIO ERROR | P:{total_pass} F:{total_fail} S:{total_skip} E:{total_error}"
            )
            if error_text:
                lines.append("")
                lines.append(error_text)
            QMessageBox.critical(self, "Scenario Run ERROR", "\n".join(lines))
        else:
            self.lbl_status.setText(
                f"SCENARIO DONE | P:{total_pass} F:{total_fail} S:{total_skip} E:{total_error}"
            )
            QMessageBox.information(self, "Scenario Run finished", "\n".join(lines))

        self.on_refresh_runs()

    def on_stop_run(self):
        stopped = False

        # 1) 일반 단일 Run stop
        if self._worker and self._worker.isRunning():
            self._worker.request_stop()
            run_short = self._last_run_id[:8] if self._last_run_id else "--------"
            self.lbl_status.setText(
                f"STOPPING {run_short} | "
                f"P:{self._run_pass_count} F:{self._run_fail_count} "
                f"S:{self._run_skip_count} E:{self._run_error_count}"
            )
            stopped = True

        # 2) Scenario Run stop
        if self._scenario_worker and self._scenario_worker.isRunning():
            self._scenario_worker.request_stop()
            self.lbl_status.setText(
                f"SCENARIO STOPPING | "
                f"{self._scenario_processed_cases}/{self._scenario_total_cases}"
            )
            stopped = True

        # 3) progress bar 문구
        if stopped:
            self.progress_run.setFormat("Stopping...")

    def on_create_rerun(self):
        if not self._last_run_id:
            QMessageBox.information(self, "No run", "Run first to generate FAIL-based re-run.")
            return
        if not self._current_plan_node_id:
            return
        ctx = self._plans[self._current_plan_node_id]

        try:
            new_preset_id = self.svc.create_rerun_preset_from_fail(
                project_id=ctx.project_id,
                base_preset_id=ctx.preset_id,
                run_id=self._last_run_id
            )
            QMessageBox.information(self, "Re-run preset created", f"New preset created.\nPreset ID: {new_preset_id}")
            # preset 콤보 갱신
            self._reload_presets(ctx.project_id, select_preset_id=new_preset_id)
        except Exception as e:
            QMessageBox.warning(self, "Re-run failed", str(e))
    
       
    def on_refresh_runs(self):
        if not self.project_id:
            QMessageBox.warning(self, "No project", "Select a project.")
            return

        try:
            runs = self.svc.list_runs_for_results(self.project_id, limit=100)
        except Exception as e:
            QMessageBox.critical(self, "Load Runs Failed", str(e))
            return

        self.run_combo.clear()
        for r in runs:
            run_id = r.get("run_id", "")
            started_at = r.get("started_at", "")
            status = r.get("status", "")
            preset_name = r.get("preset_name", "") or r.get("preset_id", "")

            short_run = run_id[:8] if run_id else ""
            label = f"{started_at} | {status} | {preset_name}"

            self.run_combo.addItem(label, run_id)
    def on_load_results(self):
        run_id = self.run_combo.currentData()
        if not run_id:
            QMessageBox.warning(self, "No run", "Select a run.")
            return

        status_filter = self.result_filter_status.currentText()

        try:
            rows = self.svc.get_results_page(
                project_id=self.project_id,
                run_id=run_id,
                status_filter=status_filter,
                offset=0,
                limit=500,
            )
        except Exception as e:
            QMessageBox.critical(self, "Load Results Failed", str(e))
            return

        # 원본 보관
        self._last_results_rows = rows

        # 현재 결과를 기준으로 필터 옵션 동적 구성
        self._refresh_result_filter_options(rows)

        test_type_filter = self.result_filter_test_type.currentText()
        band_filter = self.result_filter_band.currentText()
        standard_filter = self.result_filter_standard.currentText()
        bw_filter = self.result_filter_bw.currentText()
        channel_filter = self.result_filter_channel.currentText()
        search_text = self.result_search.text().strip().lower()

        filtered = []
        for r in rows:
            if test_type_filter != "ALL" and str(r.get("test_type", "")) != test_type_filter:
                continue

            if band_filter != "ALL" and str(r.get("band", "")) != band_filter:
                continue

            if standard_filter != "ALL" and str(r.get("standard", "")) != standard_filter:
                continue

            if bw_filter != "ALL" and str(r.get("bw_mhz", "")) != bw_filter:
                continue

            if channel_filter != "ALL" and str(r.get("channel", "")) != channel_filter:
                continue

            if search_text:
                hay = " ".join([
                    str(r.get("test_type", "")),
                    str(r.get("band", "")),
                    str(r.get("standard", "")),
                    str(r.get("group", "")),
                    str(r.get("channel", "")),
                    str(r.get("bw_mhz", "")),
                    str(r.get("reason", "")),
                    str(r.get("test_key", "")),
                ]).lower()

                if search_text not in hay:
                    continue

            filtered.append(r)

        self.results_model.set_rows(filtered)
        self.results_table.resizeColumnsToContents()
        self._update_results_summary(filtered)
        self._update_result_quick_buttons_style()
        
    def on_clear_result_filters(self):
        self.result_filter_status.setCurrentText("ALL")
        self.result_filter_test_type.setCurrentText("ALL")
        self.result_filter_band.setCurrentText("ALL")
        self.result_filter_standard.setCurrentText("ALL")
        self.result_filter_bw.setCurrentText("ALL")
        self.result_filter_channel.setCurrentText("ALL")
        self.result_search.clear()
        self._update_result_quick_buttons_style()

        if self.run_combo.currentData():
            self.on_load_results()
            
    def _refresh_result_filter_options(self, rows: list[dict]):
        # 현재 선택값 유지
        current_test = self.result_filter_test_type.currentText()
        current_band = self.result_filter_band.currentText()
        current_standard = self.result_filter_standard.currentText()
        current_bw = self.result_filter_bw.currentText()
        current_channel = self.result_filter_channel.currentText()

        # rows에서 실제 존재하는 값 추출
        test_types = sorted({
            str(r.get("test_type", "")).strip()
            for r in rows
            if r.get("test_type")
        })

        bands = sorted({
            str(r.get("band", "")).strip()
            for r in rows
            if r.get("band")
        })

        standards = sorted({
            str(r.get("standard", "")).strip()
            for r in rows
            if r.get("standard")
        })

        # BW / Channel은 숫자 정렬이 보기 좋음
        bw_values = sorted({
            int(r.get("bw_mhz"))
            for r in rows
            if r.get("bw_mhz") is not None and str(r.get("bw_mhz")).strip() != ""
        })

        channel_values = sorted({
            int(r.get("channel"))
            for r in rows
            if r.get("channel") is not None and str(r.get("channel")).strip() != ""
        })

        # signal 잠시 막기
        combos = [
            self.result_filter_test_type,
            self.result_filter_band,
            self.result_filter_standard,
            self.result_filter_bw,
            self.result_filter_channel,
        ]
        for cb in combos:
            cb.blockSignals(True)

        # test_type
        self.result_filter_test_type.clear()
        self.result_filter_test_type.addItem("ALL")
        for v in test_types:
            self.result_filter_test_type.addItem(v)

        # band
        self.result_filter_band.clear()
        self.result_filter_band.addItem("ALL")
        for v in bands:
            self.result_filter_band.addItem(v)

        # standard
        self.result_filter_standard.clear()
        self.result_filter_standard.addItem("ALL")
        for v in standards:
            self.result_filter_standard.addItem(v)

        # bw
        self.result_filter_bw.clear()
        self.result_filter_bw.addItem("ALL")
        for v in bw_values:
            self.result_filter_bw.addItem(str(v))

        # channel
        self.result_filter_channel.clear()
        self.result_filter_channel.addItem("ALL")
        for v in channel_values:
            self.result_filter_channel.addItem(str(v))

        # 이전 선택 복원
        idx = self.result_filter_test_type.findText(current_test)
        self.result_filter_test_type.setCurrentIndex(idx if idx >= 0 else 0)

        idx = self.result_filter_band.findText(current_band)
        self.result_filter_band.setCurrentIndex(idx if idx >= 0 else 0)

        idx = self.result_filter_standard.findText(current_standard)
        self.result_filter_standard.setCurrentIndex(idx if idx >= 0 else 0)

        idx = self.result_filter_bw.findText(current_bw)
        self.result_filter_bw.setCurrentIndex(idx if idx >= 0 else 0)

        idx = self.result_filter_channel.findText(current_channel)
        self.result_filter_channel.setCurrentIndex(idx if idx >= 0 else 0)

        for cb in combos:
            cb.blockSignals(False)
            
    def _update_results_summary(self, rows: list[dict]):
        p = sum(1 for r in rows if (r.get("status") or "").upper() == "PASS")
        f = sum(1 for r in rows if (r.get("status") or "").upper() == "FAIL")
        s = sum(1 for r in rows if (r.get("status") or "").upper() == "SKIP")
        e = sum(1 for r in rows if (r.get("status") or "").upper() == "ERROR")

        self.lbl_result_summary.setText(
            f"PASS {p} | FAIL {f} | SKIP {s} | ERROR {e}"
        )
        
    def _update_result_quick_buttons_style(self):
        base_style = ""
        active_all = "background-color: #E8F0FE; font-weight: bold;"
        active_fail = "background-color: #FDECEC; color: #C62828; font-weight: bold;"
        active_error = "background-color: #FFE5E5; color: #8E0000; font-weight: bold;"

        self.btn_show_all_results.setStyleSheet(base_style)
        self.btn_fail_only.setStyleSheet(base_style)
        self.btn_error_only.setStyleSheet(base_style)

        status = self.result_filter_status.currentText()

        if status == "ALL":
            self.btn_show_all_results.setStyleSheet(active_all)
        elif status == "FAIL":
            self.btn_fail_only.setStyleSheet(active_fail)
        elif status == "ERROR":
            self.btn_error_only.setStyleSheet(active_error)
        
    def _fetch_results_for_export(self, limit: int = 20000) -> list[dict]:
        run_id = self.run_combo.currentData()
        if not run_id:
            raise ValueError("No run selected")

        status_filter = self.result_filter_status.currentText()

        # 현재 화면은 limit=500으로 로드하지만, export는 크게 가져오자
        return self.svc.get_results_page(
            project_id=self.project_id,
            run_id=run_id,
            status_filter=status_filter,
            offset=0,
            limit=limit,
    )
        
    def on_export_results_csv(self):
        try:
            rows = self._fetch_results_for_export(limit=20000)
        except Exception as e:
            QMessageBox.critical(self, "Export CSV Failed", str(e))
            return

        if not rows:
            QMessageBox.information(self, "Export CSV", "No rows to export.")
            return

        path, _ = QFileDialog.getSaveFileName(
            self,
            "Export Results (CSV)",
            "results.csv",
            "CSV (*.csv)"
        )
        if not path:
            return

        # 컬럼 순서(원하는대로 바꿔도 됨)
        cols = [
            ("status", "Status"),
            ("test_type", "Test"),
            ("band", "Band"),
            ("standard", "Standard"),
            ("group", "Group"),
            ("channel", "CH"),
            ("bw_mhz", "BW(MHz)"),
            ("margin_db", "Margin(dB)"),
            ("measured_value", "Measured"),
            ("limit_value", "Limit"),
            ("reason", "Reason"),
            ("test_key", "Key"),
            ("result_id", "Result ID"),
        ]

        try:
            with open(path, "w", newline="", encoding="utf-8-sig") as f:
                w = csv.writer(f)
                w.writerow([h for _, h in cols])
                for r in rows:
                    w.writerow([r.get(k, "") for k, _ in cols])
        except Exception as e:
            QMessageBox.critical(self, "Export CSV Failed", str(e))
            return

        QMessageBox.information(self, "Export CSV", f"Saved:\n{path}")
        
    def on_export_results_excel(self):
        try:
            rows = self._fetch_results_for_export(limit=20000)
        except Exception as e:
            QMessageBox.critical(self, "Export Excel Failed", str(e))
            return

        if not rows:
            QMessageBox.information(self, "Export Excel", "No rows to export.")
            return

        path, _ = QFileDialog.getSaveFileName(
            self,
            "Export Results (Excel)",
            "results.xlsx",
            "Excel (*.xlsx)"
        )
        if not path:
            return

        cols = [
            ("status", "Status"),
            ("test_type", "Test"),
            ("band", "Band"),
            ("standard", "Standard"),
            ("group", "Group"),
            ("channel", "CH"),
            ("bw_mhz", "BW(MHz)"),
            ("margin_db", "Margin(dB)"),
            ("measured_value", "Measured"),
            ("limit_value", "Limit"),
            ("reason", "Reason"),
            ("test_key", "Key"),
            ("result_id", "Result ID"),
        ]

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Results"

            # Header
            header_font = Font(bold=True)
            for c, (_, header) in enumerate(cols, start=1):
                cell = ws.cell(row=1, column=c, value=header)
                cell.font = header_font
                cell.alignment = Alignment(vertical="center")

            # Rows
            for r_i, r in enumerate(rows, start=2):
                for c_i, (k, _) in enumerate(cols, start=1):
                    ws.cell(row=r_i, column=c_i, value=r.get(k, ""))

            # Column widths (간단 자동)
            for c_i in range(1, len(cols) + 1):
                ws.column_dimensions[ws.cell(row=1, column=c_i).column_letter].width = 16
            ws.column_dimensions["K"].width = 40  # Reason (대충 넓게)

            wb.save(path)

        except Exception as e:
            QMessageBox.critical(self, "Export Excel Failed", str(e))
            return

        QMessageBox.information(self, "Export Excel", f"Saved:\n{path}")
        
    def on_rerun_from_selection(self):
        if not self.project_id:
            QMessageBox.information(self, "No project", "Select a project first.")
            return

        # base preset은 “현재 Plan 탭에서 선택된 preset”을 기준으로 하자
        base_preset_id = self.preset_id
        if not base_preset_id:
            QMessageBox.information(self, "No base preset", "Select a preset (base) first.")
            return

        sel = self.results_table.selectionModel().selectedRows()
        if not sel:
            QMessageBox.information(self, "No selection", "Select result rows first.")
            return

        selected_rows = []
        for idx in sel:
            r = self.results_model.get_row(idx.row())
            if r:
                selected_rows.append(r)

        try:
            new_preset_id = self.svc.create_rerun_preset_from_selected_results(
                project_id=self.project_id,
                base_preset_id=base_preset_id,
                selected_rows=selected_rows
            )
            QMessageBox.information(self, "Re-run preset created", f"New preset created.\nPreset ID: {new_preset_id}")

            # preset 콤보 갱신해서 바로 선택되게
            self._reload_presets(self.project_id, select_preset_id=new_preset_id)

        except Exception as e:
            QMessageBox.warning(self, "Re-run failed", str(e))
            
    def on_edit_execution_order(self):
        if not self.preset_id:
            QMessageBox.information(self, "No preset", "Select a preset first.")
            return

        # preset json에서 기존 order 읽기
        pj = self.svc.repo.load_preset(self.preset_id)
        if "selection" in pj:
            sel = pj.get("selection", {})
        else:
            sel = pj  # 구포맷 fallback

        pol = sel.get("execution_policy", {})
        current_order = pol.get("test_order") or ["PSD", "OBW", "SP", "RX"]

        dlg = ExecutionOrderDialog(initial_order=current_order, parent=self)
        if dlg.exec() == dlg.Accepted:
            new_order = dlg.get_order()
            self.svc.save_execution_order(self.preset_id, new_order)
            QMessageBox.information(self, "Saved", f"Execution order saved:\n{new_order}")

            # Plan을 이미 올려놨다면 reload해서 반영 (선택)
            if self._current_plan_node_id:
                self.on_reload_plan()
                
    def on_result_selection_changed(self, selected, deselected):
        if not self.project_id:
            return
        sel = self.results_table.selectionModel().selectedRows()
        if not sel:
            self.steps_model.set_rows([])
            return

        row = self.results_model.get_row(sel[0].row())
        result_id = row.get("result_id")
        if not result_id:
            self.steps_model.set_rows([])
            return

        steps = self.run_repo.list_step_results(self.project_id, result_id)
        self.steps_model.set_rows(steps)
        self.steps_table.resizeColumnsToContents()
        
    def _scenario_plan_ids_in_tree_order(self) -> list[str]:
        root = self.tree_model.invisibleRootItem()
        out: list[str] = []
        for row in range(root.rowCount()):
            it = root.child(row)
            if not it:
                continue
            plan_id = it.data(Qt.UserRole)
            if plan_id:
                out.append(str(plan_id))
        return out
    
    def _clear_scenario_internal(self):
        # 1) 메모리 컨텍스트 제거
        self._plans.clear()

        # 2) 트리 제거
        root = self.tree_model.invisibleRootItem()
        if root.rowCount() > 0:
            root.removeRows(0, root.rowCount())

        # 3) 선택/페이지 상태 초기화 (네 변수명에 맞춰 최소한만)
        if hasattr(self, "_current_plan_node_id"):
            self._current_plan_node_id = None
        if hasattr(self, "_current_filter"):
            self._current_filter = None
        if hasattr(self, "_current_offset"):
            self._current_offset = 0
        if hasattr(self, "case_model"):
            try:
                self.case_model.clear()
            except Exception:
                pass        
            
    def on_save_scenario(self):
        if not self.project_id:
            QMessageBox.warning(self, "No project", "Select a project.")
            return

        plan_ids = self._scenario_plan_ids_in_tree_order()
        plans = []
        for pid in plan_ids:
            ctx = self._plans.get(pid)
            if not ctx:
                continue
            plans.append({"plan_id": pid, "preset_id": ctx.preset_id})

        data = {
            "version": "1.0",
            "saved_at": datetime.now().isoformat(timespec="seconds"),
            "project_id": self.project_id,
            "plans": plans,
        }

        path, _ = QFileDialog.getSaveFileName(
            self,
            "Save Scenario",
            "scenario.json",
            "Scenario (*.json)"
        )
        if not path:
            return

        try:
            with open(path, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
        except Exception as e:
            QMessageBox.critical(self, "Save failed", str(e))
            return

        QMessageBox.information(self, "Saved", f"Scenario saved:\n{path}")
        
    def on_load_scenario(self):
        path, _ = QFileDialog.getOpenFileName(
            self,
            "Load Scenario",
            "",
            "Scenario (*.json)"
        )
        if not path:
            return

        try:
            with open(path, "r", encoding="utf-8") as f:
                data = json.load(f)
        except Exception as e:
            QMessageBox.critical(self, "Load failed", str(e))
            return

        file_project_id = data.get("project_id")
        plans = data.get("plans", [])

        if not self.project_id:
            QMessageBox.warning(self, "No project", "Select a project before loading.")
            return

        # 프로젝트가 다르면 경고만 (자동 변경은 위험해서 권장 X)
        if file_project_id and file_project_id != self.project_id:
            ret = QMessageBox.question(
                self,
                "Project mismatch",
                f"Scenario project_id differs.\n\n"
                f"File: {file_project_id}\nCurrent: {self.project_id}\n\n"
                f"Load anyway?",
                QMessageBox.Yes | QMessageBox.No
            )
            if ret != QMessageBox.Yes:
                return

        # 1) 기존 scenario clear
        self._clear_scenario_internal()

        # 2) plan rebuild
        failures: list[str] = []
        for p in plans:
            preset_id = p.get("preset_id")
            if not preset_id:
                continue

            try:
                ruleset, preset, recipe, overrides = self.svc.build_recipe_from_preset(preset_id)
            except Exception as e:
                failures.append(f"{preset_id}: {e}")
                continue

            # file의 plan_id를 쓰되, 없으면 새로 생성
            plan_id = p.get("plan_id") or f"PLAN::{uuid.uuid4()}"

            ctx = PlanContext(
                project_id=self.project_id,
                preset_id=preset_id,
                ruleset=ruleset,
                preset=preset,
                recipe=recipe,
                overrides=overrides
            )
            self._plans[str(plan_id)] = ctx

            # 트리에 추가 (on_add_plan과 동일한 렌더링)
            root = self.tree_model.invisibleRootItem()
            plan_item = QStandardItem(f"{preset.name}  ({recipe.band}/{recipe.standard}/{recipe.plan_mode})")
            plan_item.setData(str(plan_id), role=Qt.UserRole)
            root.appendRow(plan_item)

            for t in recipe.test_types:
                child = QStandardItem(f"{t}")
                child.setData(str(plan_id), role=Qt.UserRole)
                child.setData({"test_type": t}, role=Qt.UserRole + 1)
                plan_item.appendRow(child)

        # 3) UI 마무리: 첫 plan 선택
        root = self.tree_model.invisibleRootItem()
        if root.rowCount() > 0:
            first = root.child(0)
            if first:
                self.tree.expand(first.index())
                self.tree.setCurrentIndex(first.index())
                self._select_tree_node(first)

        if failures:
            QMessageBox.warning(
                self,
                "Loaded with warnings",
                "Some plans failed to load:\n\n" + "\n".join(failures[:30])
            )
        else:
            QMessageBox.information(self, "Loaded", "Scenario loaded.")
            
    def on_clear_scenario(self):
        if not self._plans:
            return

        ret = QMessageBox.question(
            self,
            "Clear Scenario",
            "Remove all plans from the scenario?\n(This does NOT delete presets/runs.)",
            QMessageBox.Yes | QMessageBox.No
        )
        if ret != QMessageBox.Yes:
            return

        self._clear_scenario_internal()
        
    def on_results_show_all(self):
        self.result_filter_status.setCurrentText("ALL")
        self._update_result_quick_buttons_style()
        self.on_load_results()

    def on_results_fail_only(self):
        self.result_filter_status.setCurrentText("FAIL")
        self._update_result_quick_buttons_style()
        self.on_load_results()

    def on_results_error_only(self):
        self.result_filter_status.setCurrentText("ERROR")
        self._update_result_quick_buttons_style()
        self.on_load_results()
                    
class RunWorker(QThread):
    progress = Signal(int, str)   # count, last_status
    finished = Signal(str, str, str)   # final_status, run_id, error_text

    def __init__(self, run_service, project_id, preset_id, run_id, ruleset, recipe, overrides):
        super().__init__()
        self.run_service = run_service
        self.project_id = project_id
        self.preset_id = preset_id
        self.run_id = run_id
        self.ruleset = ruleset
        self.recipe = recipe
        self.overrides = overrides
        self._stop = False
        self._error_text = ""

    def request_stop(self):
        self._stop = True

    def run(self):
        import traceback
        self._error_text = ""

        try:
            def should_stop():
                return self._stop

            def on_progress(count, status):
                self.progress.emit(count, status)

            final_status = self.run_service.run(
                project_id=self.project_id,
                preset_id=self.preset_id,
                run_id=self.run_id,
                ruleset=self.ruleset,
                recipe=self.recipe,
                overrides=self.overrides,
                should_stop=should_stop,
                on_progress=on_progress,
            )

        except Exception:
            self._error_text = traceback.format_exc()
            final_status = "ERROR"

        # ✅ 여기 중요: 항상 finished emit
        self.finished.emit(final_status, self.run_id, self._error_text)


#-----------------Scenario를 실행하는 Class-------------------------        
class ScenarioRunWorker(QThread):
    progress = Signal(int, int, str, str)   # processed, total, preset_name, last_status
    finished = Signal(str, object, str)     # final_status, summaries(list), error_text

    def __init__(self, run_service, run_repo, plan_snapshots, total_cases: int):
        super().__init__()
        self.run_service = run_service
        self.run_repo = run_repo
        self.plan_snapshots = plan_snapshots
        self.total_cases = total_cases
        self._stop = False
        self._error_text = ""

    def request_stop(self):
        self._stop = True

    def run(self):
        import traceback

        summaries = []
        processed_global = 0

        try:
            for plan in self.plan_snapshots:
                if self._stop:
                    break

                project_id = plan["project_id"]
                preset_id = plan["preset_id"]
                preset_name = plan["preset_name"]
                ruleset = plan["ruleset"]
                recipe = plan["recipe"]
                overrides = plan["overrides"]

                run_id = self.run_repo.create_run(project_id, preset_id)

                def should_stop():
                    return self._stop

                def on_progress(count, status):
                    nonlocal processed_global
                    # count는 "현재 preset 내 count"
                    base = sum(p.get("case_count", 0) for p in summaries)
                    processed_global = base + count
                    self.progress.emit(processed_global, self.total_cases, preset_name, status)

                final_status = self.run_service.run(
                    project_id=project_id,
                    preset_id=preset_id,
                    run_id=run_id,
                    ruleset=ruleset,
                    recipe=recipe,
                    overrides=overrides,
                    should_stop=should_stop,
                    on_progress=on_progress,
                )

                self.run_repo.finish_run(run_id, final_status)

                try:
                    counts = self.run_repo.get_run_status_counts(project_id, run_id)
                except Exception:
                    counts = {"PASS": 0, "FAIL": 0, "SKIP": 0, "ERROR": 0}

                summaries.append({
                    "plan_id": plan["plan_id"],
                    "preset_id": preset_id,
                    "preset_name": preset_name,
                    "run_id": run_id,
                    "final_status": final_status,
                    "case_count": plan.get("case_count", 0),
                    "counts": counts,
                })

            final = "STOPPED" if self._stop else "DONE"
            self.finished.emit(final, summaries, self._error_text)

        except Exception:
            self._error_text = traceback.format_exc()
            self.finished.emit("ERROR", summaries, self._error_text)