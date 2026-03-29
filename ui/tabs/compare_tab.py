from __future__ import annotations

import csv
import logging
from typing import Callable, Dict, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from PySide6.QtGui import QColor, QBrush, QFont, QStandardItem, QStandardItemModel
from PySide6.QtWidgets import (
    QCheckBox,
    QComboBox,
    QFileDialog,
    QHBoxLayout,
    QLabel,
    QMessageBox,
    QPushButton,
    QTableView,
    QVBoxLayout,
    QWidget,
)

from ui.workers.results_task_worker import ResultsTaskWorker


log = logging.getLogger(__name__)


class CompareTab(QWidget):
    def __init__(self, service, get_project_id: Callable[[], str | None], parent=None):
        super().__init__(parent)
        self.svc = service
        self.get_project_id = get_project_id
        self._last_compare_rows: List[Dict] = []
        self._task_worker: ResultsTaskWorker | None = None
        self._task_generation = 0
        self._busy_action = ""
        self._build_ui()

    def refresh_runs(self) -> None:
        self.on_refresh_compare_runs()

    def clear_compare(self) -> None:
        self._last_compare_rows = []
        self.compare_model.removeRows(0, self.compare_model.rowCount())
        self.lbl_compare_summary.setText("No comparison loaded")

    def reset_view(self) -> None:
        self._task_generation += 1
        self._task_worker = None
        self.compare_run_a.blockSignals(True)
        self.compare_run_b.blockSignals(True)
        self.compare_run_a.clear()
        self.compare_run_b.clear()
        self.compare_run_a.blockSignals(False)
        self.compare_run_b.blockSignals(False)
        self.chk_compare_changes_only.blockSignals(True)
        self.chk_compare_changes_only.setChecked(False)
        self.chk_compare_changes_only.blockSignals(False)
        self.clear_compare()
        self._set_busy(False)

    def _build_ui(self) -> None:
        layout = QVBoxLayout(self)

        top = QHBoxLayout()
        self.compare_run_a = QComboBox()
        self.compare_run_b = QComboBox()
        self.btn_refresh_compare_runs = QPushButton("Refresh Runs")
        self.btn_load_compare = QPushButton("Compare")
        self.chk_compare_changes_only = QCheckBox("Changed only")
        self.lbl_compare_summary = QLabel("No comparison loaded")

        top.addWidget(QLabel("Run A:"))
        top.addWidget(self.compare_run_a, 2)
        top.addSpacing(8)
        top.addWidget(QLabel("Run B:"))
        top.addWidget(self.compare_run_b, 2)
        top.addWidget(self.chk_compare_changes_only)
        top.addWidget(self.btn_refresh_compare_runs)
        top.addWidget(self.btn_load_compare)
        layout.addLayout(top)

        actions = QHBoxLayout()
        self.btn_export_compare_csv = QPushButton("Export CSV")
        self.btn_export_compare_excel = QPushButton("Export Excel")
        actions.addWidget(self.btn_export_compare_csv)
        actions.addWidget(self.btn_export_compare_excel)
        actions.addStretch(1)
        actions.addWidget(self.lbl_compare_summary)
        layout.addLayout(actions)

        self.compare_table = QTableView()
        self.compare_model = QStandardItemModel()
        self.compare_model.setHorizontalHeaderLabels([
            "Test", "Band", "Std", "CH", "BW",
            "Run A", "Run B", "Margin A", "Margin B", "Δ Margin", "Changed"
        ])
        self.compare_table.setModel(self.compare_model)
        self.compare_table.setSelectionBehavior(QTableView.SelectRows)
        self.compare_table.setSelectionMode(QTableView.ExtendedSelection)
        self.compare_table.horizontalHeader().setStretchLastSection(True)
        self.compare_table.setSortingEnabled(True)
        self.compare_table.setAlternatingRowColors(True)
        self.compare_table.setStyleSheet("QTableView { alternate-background-color: #141b24; gridline-color: #334155; selection-background-color: #1d4ed8; selection-color: white; }")
        layout.addWidget(self.compare_table, 1)

        self.btn_refresh_compare_runs.clicked.connect(self.on_refresh_compare_runs)
        self.btn_load_compare.clicked.connect(self.on_load_compare)
        self.chk_compare_changes_only.stateChanged.connect(self.on_load_compare)
        self.btn_export_compare_csv.clicked.connect(self.on_export_compare_csv)
        self.btn_export_compare_excel.clicked.connect(self.on_export_compare_excel)

    def _set_busy(self, busy: bool, action: str = "") -> None:
        self._busy_action = action if busy else ""
        self.btn_refresh_compare_runs.setEnabled(not busy)
        self.btn_load_compare.setEnabled(not busy)
        self.btn_export_compare_csv.setEnabled(not busy)
        self.btn_export_compare_excel.setEnabled(not busy)
        self.chk_compare_changes_only.setEnabled(not busy)

    def _finish_worker(self, worker: ResultsTaskWorker) -> None:
        if self._task_worker is worker:
            self._task_worker = None
        worker.deleteLater()
        self._set_busy(False)

    def _start_task(self, *, action: str, task, on_success, error_title: str) -> None:
        if self._task_worker is not None and self._task_worker.isRunning():
            log.info("compare task skipped | busy_action=%s next_action=%s", self._busy_action, action)
            return

        self._task_generation += 1
        generation = self._task_generation
        worker = ResultsTaskWorker(task)
        self._task_worker = worker
        self._set_busy(True, action=action)

        def _handle_success(payload) -> None:
            self._finish_worker(worker)
            if generation != self._task_generation:
                return
            on_success(payload)

        def _handle_failure(error_text: str) -> None:
            self._finish_worker(worker)
            if generation != self._task_generation:
                return
            log.error("compare task failed | action=%s\n%s", action, error_text)
            QMessageBox.critical(self, error_title, error_text)

        worker.succeeded.connect(_handle_success)
        worker.failed.connect(_handle_failure)
        worker.start()

    def _fill_run_combo(self, combo: QComboBox, runs: List[Dict]) -> None:
        current = combo.currentData()
        combo.blockSignals(True)
        combo.clear()
        for r in runs:
            run_id = r.get("run_id", "")
            started_at = r.get("started_at", "")
            status = r.get("status", "")
            preset_name = r.get("preset_name", "") or r.get("preset_id", "")
            combo.addItem(f"{started_at} | {status} | {preset_name}", run_id)

        if current:
            idx = combo.findData(current)
            if idx >= 0:
                combo.setCurrentIndex(idx)
        combo.blockSignals(False)

    def on_refresh_compare_runs(self):
        project_id = self.get_project_id()
        if not project_id:
            self.reset_view()
            return
        try:
            runs = self.svc.list_runs_for_results(project_id, limit=100)
        except Exception as e:
            QMessageBox.critical(self, "Load Runs Failed", str(e))
            return

        self._fill_run_combo(self.compare_run_a, runs)
        self._fill_run_combo(self.compare_run_b, runs)
        if self.compare_run_b.count() > 1 and self.compare_run_b.currentIndex() == 0:
            self.compare_run_b.setCurrentIndex(1)

    def _render_compare_rows(self, rows: List[Dict]) -> None:
        headers = ["Test", "Band", "Std", "CH", "BW", "Run A", "Run B", "Margin A", "Margin B", "Δ Margin", "Changed"]
        self.compare_model.clear()
        self.compare_model.setHorizontalHeaderLabels(headers)

        for row in rows:
            items = [
                QStandardItem(str(row.get("test_type", ""))),
                QStandardItem(str(row.get("band", ""))),
                QStandardItem(str(row.get("standard", ""))),
                QStandardItem(str(row.get("channel", ""))),
                QStandardItem(str(row.get("bw_mhz", ""))),
                QStandardItem(str(row.get("status_a", ""))),
                QStandardItem(str(row.get("status_b", ""))),
                QStandardItem(str(row.get("margin_a", ""))),
                QStandardItem(str(row.get("margin_b", ""))),
                QStandardItem(str(row.get("delta_margin", ""))),
                QStandardItem("Y" if row.get("changed") else ""),
            ]

            status_a = str(row.get("status_a", ""))
            status_b = str(row.get("status_b", ""))
            delta = row.get("delta_margin", "")
            changed = bool(row.get("changed"))

            row_color = None
            text_color = None
            status_font = QFont()
            if status_a == "PASS" and status_b == "FAIL":
                row_color = QColor(123, 31, 31)      # deeper muted red
                text_color = QColor(255, 245, 245)
                status_font.setBold(True)
            elif status_a == "FAIL" and status_b == "PASS":
                row_color = QColor(27, 94, 32)       # deeper muted green
                text_color = QColor(245, 255, 245)
                status_font.setBold(True)
            elif "MISSING" in (status_a, status_b):
                row_color = QColor(55, 65, 81)       # slate gray
                text_color = QColor(248, 250, 252)
            elif changed:
                try:
                    if abs(float(delta)) >= 3:
                        row_color = QColor(120, 53, 15)   # strong amber/brown
                        text_color = QColor(255, 251, 235)
                    else:
                        row_color = QColor(92, 25, 27)    # muted rose for subtle change
                        text_color = QColor(255, 241, 242)
                except Exception:
                    row_color = QColor(92, 25, 27)
                    text_color = QColor(255, 241, 242)

            if row_color is not None:
                for idx, item in enumerate(items):
                    item.setBackground(row_color)
                    if text_color is not None:
                        item.setForeground(QBrush(text_color))
                    if idx in (5, 6, 9, 10):
                        item.setFont(status_font)

            self.compare_model.appendRow(items)

        self.compare_table.resizeColumnsToContents()

    def on_load_compare(self):
        project_id = self.get_project_id()
        if not project_id:
            return

        run_a = self.compare_run_a.currentData()
        run_b = self.compare_run_b.currentData()
        if not run_a or not run_b:
            return
        if run_a == run_b:
            self.lbl_compare_summary.setText("Select two different runs.")
            self.compare_model.removeRows(0, self.compare_model.rowCount())
            return

        changed_only = self.chk_compare_changes_only.isChecked()

        def _task():
            rows = self.svc.get_comparable_results(project_id, run_a, run_b)
            if changed_only:
                rows = [r for r in rows if r.get("changed")]
            return {
                "project_id": project_id,
                "run_a": run_a,
                "run_b": run_b,
                "changed_only": changed_only,
                "rows": rows,
            }

        def _apply(payload: Dict) -> None:
            if payload.get("project_id") != self.get_project_id():
                return
            if payload.get("run_a") != self.compare_run_a.currentData():
                return
            if payload.get("run_b") != self.compare_run_b.currentData():
                return
            if bool(payload.get("changed_only")) != self.chk_compare_changes_only.isChecked():
                return
            rows = list(payload.get("rows") or [])
            self._last_compare_rows = rows
            changed_count = sum(1 for r in rows if r.get("changed"))
            self.lbl_compare_summary.setText(f"Rows {len(rows)} | Changed {changed_count}")
            self._render_compare_rows(rows)

        self._start_task(
            action="load_compare",
            task=_task,
            on_success=_apply,
            error_title="Compare Failed",
        )

    def _fetch_compare_rows_for_export(self) -> List[Dict]:
        if not self._last_compare_rows:
            raise ValueError("No compare rows loaded")
        return list(self._last_compare_rows)

    def on_export_compare_csv(self):
        try:
            rows = self._fetch_compare_rows_for_export()
        except Exception as e:
            QMessageBox.critical(self, "Export Compare CSV Failed", str(e))
            return

        path, _ = QFileDialog.getSaveFileName(self, "Export Compare (CSV)", "compare_results.csv", "CSV (*.csv)")
        if not path:
            return

        cols = [
            ("test_type", "Test"),
            ("band", "Band"),
            ("standard", "Standard"),
            ("channel", "CH"),
            ("bw_mhz", "BW(MHz)"),
            ("status_a", "Run A Status"),
            ("status_b", "Run B Status"),
            ("margin_a", "Margin A"),
            ("margin_b", "Margin B"),
            ("delta_margin", "Delta Margin"),
            ("changed", "Changed"),
        ]
        def _task():
            with open(path, "w", newline="", encoding="utf-8-sig") as f:
                w = csv.writer(f)
                w.writerow([h for _, h in cols])
                for r in rows:
                    w.writerow([r.get(k, "") for k, _ in cols])
            return {"path": path}

        def _done(payload: Dict) -> None:
            QMessageBox.information(self, "Export Compare CSV", f"Saved:\n{payload.get('path', path)}")

        self._start_task(
            action="export_compare_csv",
            task=_task,
            on_success=_done,
            error_title="Export Compare CSV Failed",
        )

    def on_export_compare_excel(self):
        try:
            rows = self._fetch_compare_rows_for_export()
        except Exception as e:
            QMessageBox.critical(self, "Export Compare Excel Failed", str(e))
            return

        path, _ = QFileDialog.getSaveFileName(self, "Export Compare (Excel)", "compare_results.xlsx", "Excel (*.xlsx)")
        if not path:
            return

        cols = [
            ("test_type", "Test"),
            ("band", "Band"),
            ("standard", "Standard"),
            ("channel", "CH"),
            ("bw_mhz", "BW(MHz)"),
            ("status_a", "Run A Status"),
            ("status_b", "Run B Status"),
            ("margin_a", "Margin A"),
            ("margin_b", "Margin B"),
            ("delta_margin", "Delta Margin"),
            ("changed", "Changed"),
        ]

        red_fill = PatternFill(fill_type="solid", fgColor="7B1F1F")
        green_fill = PatternFill(fill_type="solid", fgColor="1B5E20")
        amber_fill = PatternFill(fill_type="solid", fgColor="78350F")
        yellow_fill = PatternFill(fill_type="solid", fgColor="5C191B")
        gray_fill = PatternFill(fill_type="solid", fgColor="374151")
        header_font = Font(bold=True)
        light_font = Font(color="FFFFFF")

        def _task():
            wb = Workbook()
            ws = wb.active
            ws.title = "Compare"

            for c, (_, header) in enumerate(cols, start=1):
                cell = ws.cell(row=1, column=c, value=header)
                cell.font = header_font
                cell.alignment = Alignment(vertical="center")

            for r_i, r in enumerate(rows, start=2):
                for c_i, (k, _) in enumerate(cols, start=1):
                    value = "Y" if (k == "changed" and r.get("changed")) else r.get(k, "")
                    ws.cell(row=r_i, column=c_i, value=value)

                status_a = str(r.get("status_a", ""))
                status_b = str(r.get("status_b", ""))
                delta = r.get("delta_margin", "")
                changed = bool(r.get("changed"))

                fill = None
                if status_a == "PASS" and status_b == "FAIL":
                    fill = red_fill
                elif status_a == "FAIL" and status_b == "PASS":
                    fill = green_fill
                elif "MISSING" in (status_a, status_b):
                    fill = gray_fill
                elif changed:
                    try:
                        fill = amber_fill if abs(float(delta)) >= 3 else yellow_fill
                    except Exception:
                        fill = yellow_fill

                if fill is not None:
                    for c_i in range(1, len(cols) + 1):
                        ws.cell(row=r_i, column=c_i).fill = fill
                        ws.cell(row=r_i, column=c_i).font = light_font

            widths = [18, 10, 14, 10, 10, 14, 14, 12, 12, 14, 10]
            for c_i, width in enumerate(widths, start=1):
                ws.column_dimensions[ws.cell(row=1, column=c_i).column_letter].width = width

            wb.save(path)
            return {"path": path}

        def _done(payload: Dict) -> None:
            QMessageBox.information(self, "Export Compare Excel", f"Saved:\n{payload.get('path', path)}")

        self._start_task(
            action="export_compare_excel",
            task=_task,
            on_success=_done,
            error_title="Export Compare Excel Failed",
        )
