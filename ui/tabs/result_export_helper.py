from __future__ import annotations

import csv
from typing import Any, Dict, Iterable, List, Sequence, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from application.result_difference import format_difference, format_difference_value, format_numeric_value


ExportColumn = Tuple[str, str]


class ResultExportHelper:
    def results_export_columns(self) -> List[ExportColumn]:
        return [
            ("status", "Status"),
            ("test_type", "Test"),
            ("band", "Band"),
            ("standard", "Standard"),
            ("data_rate", "Data Rate"),
            ("bw_mhz", "BW(MHz)"),
            ("channel", "CH"),
            ("voltage_condition", "Voltage Cond"),
            ("target_voltage_v_display", "Voltage (V)"),
            ("measured_value", "Measured (Corrected)"),
            ("raw_measured_value", "Raw Measured"),
            ("applied_correction_db", "Applied Correction(dB)"),
            ("correction_profile_name", "Correction Profile"),
            ("correction_mode", "Correction Mode"),
            ("correction_bound_path", "Correction Path"),
            ("correction_breakdown_text", "Correction Breakdown"),
            ("limit_value", "Limit"),
            ("difference_display", "Difference"),
            ("measurement_unit", "Unit"),
            ("screenshot", "Screenshot"),
            ("reason", "Reason"),
            ("test_key", "Key"),
            ("result_id", "Result ID"),
        ]

    def build_results_export_row(self, row: Dict[str, Any]) -> Dict[str, Any]:
        export_row = dict(row or {})
        export_row["measured_value"] = format_numeric_value(export_row.get("measured_value"))
        export_row["raw_measured_value"] = format_numeric_value(export_row.get("raw_measured_value"))
        export_row["applied_correction_db"] = format_numeric_value(export_row.get("applied_correction_db"))
        export_row["limit_value"] = format_numeric_value(export_row.get("limit_value"))
        export_row["difference_display"] = format_difference(
            export_row.get("difference_value"),
            export_row.get("difference_unit", ""),
        )
        export_row["target_voltage_v_display"] = self.format_voltage(export_row.get("target_voltage_v"))
        export_row["screenshot"] = "Yes" if export_row.get("has_screenshot") else ""
        export_row["correction_breakdown_text"] = self.format_correction_breakdown_text(
            export_row.get("correction_breakdown") or {},
            export_row.get("applied_correction_db"),
        )
        return export_row

    def write_results_csv(self, path: str, rows: Sequence[Dict[str, Any]]) -> None:
        cols = self.results_export_columns()
        with open(path, "w", newline="", encoding="utf-8-sig") as handle:
            writer = csv.writer(handle)
            writer.writerow([header for _, header in cols])
            for row in rows:
                export_row = self.build_results_export_row(dict(row or {}))
                writer.writerow([export_row.get(key, "") for key, _ in cols])

    def write_results_excel(self, path: str, rows: Sequence[Dict[str, Any]]) -> None:
        cols = self.results_export_columns()
        wb = Workbook()
        ws = wb.active
        ws.title = "Results"
        header_font = Font(bold=True)
        for column_index, (_, header) in enumerate(cols, start=1):
            cell = ws.cell(row=1, column=column_index, value=header)
            cell.font = header_font
            cell.alignment = Alignment(vertical="center")

        for row_index, row in enumerate(rows, start=2):
            export_row = self.build_results_export_row(dict(row or {}))
            for column_index, (key, _) in enumerate(cols, start=1):
                ws.cell(row=row_index, column=column_index, value=export_row.get(key, ""))

        widths = [12, 12, 10, 14, 12, 10, 8, 16, 14, 18, 18, 18, 16, 16, 16, 40, 18, 14, 12, 12, 40, 26, 26]
        self._apply_widths(ws, widths)
        wb.save(path)

    def compare_export_columns(self) -> List[ExportColumn]:
        return [
            ("test_type", "Test"),
            ("band", "Band"),
            ("standard", "Standard"),
            ("data_rate", "Data Rate"),
            ("bw_mhz", "BW(MHz)"),
            ("channel", "CH"),
            ("voltage_condition", "Voltage Cond"),
            ("target_voltage_v_display", "Voltage (V)"),
            ("measured_a", "Run A (Corrected)"),
            ("measured_b", "Run B (Corrected)"),
            ("raw_measured_a", "Run A Raw"),
            ("raw_measured_b", "Run B Raw"),
            ("applied_correction_db_a", "Correction A(dB)"),
            ("applied_correction_db_b", "Correction B(dB)"),
            ("delta_applied_correction_db", "Correction Delta(B-A)"),
            ("correction_profile_name_a", "Correction Profile A"),
            ("correction_profile_name_b", "Correction Profile B"),
            ("correction_mode_a", "Correction Mode A"),
            ("correction_mode_b", "Correction Mode B"),
            ("correction_bound_path_a", "Correction Path A"),
            ("correction_bound_path_b", "Correction Path B"),
            ("correction_breakdown_text_a", "Correction Breakdown A"),
            ("correction_breakdown_text_b", "Correction Breakdown B"),
            ("delta_display", "Delta"),
            ("unit", "Unit"),
            ("status_a", "Status A"),
            ("status_b", "Status B"),
            ("difference_a_display", "Difference A"),
            ("difference_b_display", "Difference B"),
            ("limit_a", "Limit A"),
            ("limit_b", "Limit B"),
            ("comparator_a", "Comparator A"),
            ("comparator_b", "Comparator B"),
            ("screenshot_a", "Screenshot A"),
            ("screenshot_b", "Screenshot B"),
        ]

    def build_compare_export_row(self, row: Dict[str, Any]) -> Dict[str, Any]:
        export_row = dict(row or {})
        export_row["measured_a"] = format_numeric_value(export_row.get("measured_a"))
        export_row["measured_b"] = format_numeric_value(export_row.get("measured_b"))
        export_row["raw_measured_a"] = format_numeric_value(export_row.get("raw_measured_a"))
        export_row["raw_measured_b"] = format_numeric_value(export_row.get("raw_measured_b"))
        export_row["applied_correction_db_a"] = format_numeric_value(export_row.get("applied_correction_db_a"))
        export_row["applied_correction_db_b"] = format_numeric_value(export_row.get("applied_correction_db_b"))
        export_row["delta_applied_correction_db"] = format_difference_value(export_row.get("delta_applied_correction_db"))
        export_row["target_voltage_v_display"] = self.format_voltage(
            export_row.get("target_voltage_v_a")
            if export_row.get("target_voltage_v_a") not in (None, "")
            else export_row.get("target_voltage_v_b")
        )
        export_row["delta_display"] = format_difference_value(export_row.get("delta_value"))
        export_row["difference_a_display"] = format_difference(
            export_row.get("difference_a"),
            export_row.get("difference_unit", ""),
        )
        export_row["difference_b_display"] = format_difference(
            export_row.get("difference_b"),
            export_row.get("difference_unit", ""),
        )
        export_row["correction_breakdown_text_a"] = export_row.get("correction_breakdown_text_a") or self.format_correction_breakdown_text(
            export_row.get("correction_breakdown_a") or {},
            export_row.get("applied_correction_db_a"),
        )
        export_row["correction_breakdown_text_b"] = export_row.get("correction_breakdown_text_b") or self.format_correction_breakdown_text(
            export_row.get("correction_breakdown_b") or {},
            export_row.get("applied_correction_db_b"),
        )
        export_row["screenshot_a"] = "Yes" if export_row.get("has_screenshot_a") else ""
        export_row["screenshot_b"] = "Yes" if export_row.get("has_screenshot_b") else ""
        return export_row

    def write_compare_csv(self, path: str, rows: Sequence[Dict[str, Any]]) -> None:
        cols = self.compare_export_columns()
        with open(path, "w", newline="", encoding="utf-8-sig") as handle:
            writer = csv.writer(handle)
            writer.writerow([header for _, header in cols])
            for row in rows:
                export_row = self.build_compare_export_row(dict(row or {}))
                writer.writerow([export_row.get(key, "") for key, _ in cols])

    def write_compare_excel(self, path: str, rows: Sequence[Dict[str, Any]]) -> None:
        cols = self.compare_export_columns()
        wb = Workbook()
        ws = wb.active
        ws.title = "Compare"

        red_fill = PatternFill(fill_type="solid", fgColor="7B1F1F")
        green_fill = PatternFill(fill_type="solid", fgColor="1B5E20")
        amber_fill = PatternFill(fill_type="solid", fgColor="78350F")
        gray_fill = PatternFill(fill_type="solid", fgColor="374151")
        header_font = Font(bold=True)
        light_font = Font(color="FFFFFF")

        for column_index, (_, header) in enumerate(cols, start=1):
            cell = ws.cell(row=1, column=column_index, value=header)
            cell.font = header_font
            cell.alignment = Alignment(vertical="center")

        for row_index, row in enumerate(rows, start=2):
            export_row = self.build_compare_export_row(dict(row or {}))
            for column_index, (key, _) in enumerate(cols, start=1):
                ws.cell(row=row_index, column=column_index, value=export_row.get(key, ""))

            fill = None
            status_a = str(export_row.get("status_a", ""))
            status_b = str(export_row.get("status_b", ""))
            if status_a == "PASS" and status_b == "FAIL":
                fill = red_fill
            elif status_a == "FAIL" and status_b == "PASS":
                fill = green_fill
            elif export_row.get("changed"):
                fill = amber_fill
            elif "MISSING" in (status_a, status_b):
                fill = gray_fill

            if fill is not None:
                for column_index in range(1, len(cols) + 1):
                    ws.cell(row=row_index, column=column_index).fill = fill
                    ws.cell(row=row_index, column=column_index).font = light_font

        widths = [12, 10, 14, 12, 10, 8, 16, 12, 18, 18, 18, 18, 16, 16, 18, 18, 18, 16, 16, 16, 16, 36, 36, 16, 12, 12, 12, 16, 16, 12, 12, 14, 14, 12, 12]
        self._apply_widths(ws, widths)
        wb.save(path)

    def format_voltage(self, value: Any) -> str:
        try:
            if value in (None, ""):
                return ""
            return f"{float(value):g}"
        except Exception:
            return str(value or "")

    def format_correction_breakdown_text(self, breakdown: Dict[str, Any], applied_total: Any) -> str:
        payload = dict(breakdown or {})
        if not payload:
            if applied_total in (None, "", 0, 0.0):
                return ""
            return f"total_db={format_numeric_value(applied_total)}"

        ordered_keys = [
            "cable_loss_db",
            "attenuator_db",
            "dut_cable_loss_db",
            "switchbox_loss_db",
            "divider_loss_db",
            "external_gain_db",
            "manual_offset_db",
        ]
        parts = []
        for key in ordered_keys:
            if key not in payload:
                continue
            parts.append(f"{key}={format_numeric_value(payload.get(key))}")
        parts.append(f"total_db={format_numeric_value(applied_total)}")
        return "; ".join(parts)

    def _apply_widths(self, worksheet, widths: Iterable[int]) -> None:
        for column_index, width in enumerate(widths, start=1):
            worksheet.column_dimensions[worksheet.cell(row=1, column=column_index).column_letter].width = width


__all__ = ["ResultExportHelper", "ExportColumn"]
